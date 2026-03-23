import os, json, base64, io
from datetime import datetime, date
from pathlib import Path
from functools import wraps

from flask import Flask, request, jsonify, render_template, session, redirect, send_file
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24))
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024

DATABASE_URL   = os.environ.get("DATABASE_URL", "")
APP_PASSWORD   = os.environ.get("APP_PASSWORD", "")    # regular user password
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")  # admin password (full access)
ANTHROPIC_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")
MODEL          = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")
RESEND_KEY     = os.environ.get("RESEND_API_KEY", "")
NOTIFY_EMAIL   = os.environ.get("NOTIFY_EMAIL", "johns@boomrecords.co")
APP_URL        = os.environ.get("APP_URL", "")

CATEGORIES = ["Recording","Mixing & Mastering","Music Video","Marketing",
              "Sync/Licensing","Distribution","Legal","Merch","Tour/Live","Cobrand","Other"]
PAYMENT_METHODS = ["ACH","Check","Wire","Credit Card","PayPal","Cash"]

EXTRACT_PROMPT = """Extract the following fields from this invoice or receipt.
Return ONLY valid JSON — no markdown, no extra text:
{
  "invoice_date": "MM/DD/YYYY if found, else empty string",
  "payee": "vendor or company name",
  "description": "brief description of what was invoiced (1 sentence max)",
  "category": "best match from: Recording, Mixing & Mastering, Music Video, Marketing, Sync/Licensing, Distribution, Legal, Merch, Tour/Live, Cobrand, Other",
  "invoice_number": "invoice number if present, else empty string",
  "amount": <number, 2 decimal places, no symbols, 0 if not found>,
  "payment_method": "best match from: ACH, Check, Wire, Credit Card, PayPal, Cash — or empty string"
}"""


# ── DB ────────────────────────────────────────────────────────────────────────

def get_db():
    if DATABASE_URL:
        import psycopg2
        url = DATABASE_URL.replace("postgres://","postgresql://",1)
        return psycopg2.connect(url), "pg"
    else:
        import sqlite3
        conn = sqlite3.connect(str(Path(__file__).parent/"boom.db"))
        conn.row_factory = sqlite3.Row
        return conn, "sqlite"

def init_db():
    conn, kind = get_db()
    cur = conn.cursor()
    if kind == "pg":
        cur.execute("""CREATE TABLE IF NOT EXISTS expenses (
            id SERIAL PRIMARY KEY, invoice_date DATE, payee TEXT,
            description TEXT, category TEXT, artist TEXT, song TEXT,
            invoice_number TEXT, amount NUMERIC(12,2), payment_method TEXT,
            payment_date DATE, in_quickbooks TEXT DEFAULT 'No',
            qb_entry_date DATE, uploaded_to_stem TEXT DEFAULT 'No',
            stem_upload_date DATE, notes TEXT, vendor_submitted BOOLEAN DEFAULT FALSE,
            vendor_name TEXT, vendor_email TEXT, created_at TIMESTAMP DEFAULT NOW())""")
        for col in ["song TEXT","vendor_submitted BOOLEAN DEFAULT FALSE",
                    "vendor_name TEXT","vendor_email TEXT"]:
            cur.execute(f"ALTER TABLE expenses ADD COLUMN IF NOT EXISTS {col}")
    else:
        cur.execute("""CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT, invoice_date TEXT, payee TEXT,
            description TEXT, category TEXT, artist TEXT, song TEXT,
            invoice_number TEXT, amount REAL, payment_method TEXT,
            payment_date TEXT, in_quickbooks TEXT DEFAULT 'No',
            qb_entry_date TEXT, uploaded_to_stem TEXT DEFAULT 'No',
            stem_upload_date TEXT, notes TEXT, vendor_submitted INTEGER DEFAULT 0,
            vendor_name TEXT, vendor_email TEXT,
            created_at TEXT DEFAULT (datetime('now')))""")
        for col in ["song TEXT","vendor_submitted INTEGER DEFAULT 0","vendor_name TEXT","vendor_email TEXT"]:
            try: cur.execute(f"ALTER TABLE expenses ADD COLUMN {col}")
            except: pass
    conn.commit(); conn.close()


# ── Auth ──────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if (APP_PASSWORD or ADMIN_PASSWORD) and not session.get("authenticated"):
            return redirect("/login")
        return f(*a, **kw)
    return dec

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if session.get("role") != "admin":
            return jsonify({"error": "Admin access required"}), 403
        return f(*a, **kw)
    return dec

def is_admin():
    return session.get("role") == "admin"

@app.route("/login", methods=["GET","POST"])
def login():
    err = None
    if request.method == "POST":
        pw = request.form.get("password","")
        if ADMIN_PASSWORD and pw == ADMIN_PASSWORD:
            session["authenticated"] = True
            session["role"] = "admin"
            return redirect("/")
        elif APP_PASSWORD and pw == APP_PASSWORD:
            session["authenticated"] = True
            session["role"] = "user"
            return redirect("/")
        elif not APP_PASSWORD and not ADMIN_PASSWORD:
            session["authenticated"] = True
            session["role"] = "admin"
            return redirect("/")
        else:
            err = "Incorrect password."
    return render_template("login.html", error=err)

@app.route("/logout")
def logout():
    session.clear(); return redirect("/login")


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_date(s):
    if not s: return None
    for fmt in ("%m/%d/%Y","%Y-%m-%d","%m-%d-%Y","%d/%m/%Y"):
        try: return datetime.strptime(s.strip(), fmt).date()
        except: pass
    return None

def parse_amount(v):
    try: return float(str(v).replace("$","").replace(",","").strip()) if v else None
    except: return None

def extract_fields(file_bytes, mime):
    if not ANTHROPIC_KEY: return {}
    b64 = base64.standard_b64encode(file_bytes).decode()
    content = ([{"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64}},
                {"type":"text","text":EXTRACT_PROMPT}]
               if mime == "application/pdf" else
               [{"type":"image","source":{"type":"base64","media_type":mime,"data":b64}},
                {"type":"text","text":EXTRACT_PROMPT}])
    try:
        resp = anthropic.Anthropic(api_key=ANTHROPIC_KEY).messages.create(
            model=MODEL, max_tokens=512, messages=[{"role":"user","content":content}])
        raw = resp.content[0].text.strip()
        if raw.startswith("```"): raw = raw.split("```")[1]; raw = raw[4:] if raw.startswith("json") else raw
        return json.loads(raw.strip())
    except Exception as e:
        print(f"Extraction error: {e}"); return {}

def get_unknowns(fields):
    u = []
    if not fields.get("invoice_date"): u.append("Invoice date missing")
    if not fields.get("amount") or float(fields.get("amount",0))==0: u.append("Amount not found or zero")
    if not fields.get("invoice_number"): u.append("Invoice number missing")
    if fields.get("category") in ("Other","",None): u.append("Category could not be determined")
    if not fields.get("description"): u.append("Description missing")
    return u


# ── Email ─────────────────────────────────────────────────────────────────────

def send_vendor_email(vendor_name, vendor_email, fields, unknowns):
    if not RESEND_KEY: return
    try:
        import resend; resend.api_key = RESEND_KEY
        review_url = f"{APP_URL}/ledger" if APP_URL else "#"
        amt = fields.get("amount",0)
        amt_str = f"${float(amt):,.2f}" if amt else "Unknown"
        warn = ("".join(f"<li style='color:#d97706'>⚠ {u}</li>" for u in unknowns)
                if unknowns else "")
        warn_block = (f"<div style='background:#fef3c7;border:1px solid #fcd34d;"
                      f"border-radius:8px;padding:12px 16px;margin:14px 0'>"
                      f"<strong style='color:#92400e'>Needs review:</strong>"
                      f"<ul style='margin:6px 0 0 16px;color:#92400e'>{warn}</ul></div>") if warn else ""
        def row(bg, label, val):
            bg_s = "background:#f9f9f9;" if bg else ""
            return (f"<tr><td style='padding:7px 12px;{bg_s}color:#666;width:150px'>{label}</td>"
                    f"<td style='padding:7px 12px;{bg_s}'>{val}</td></tr>")
        html = f"""<div style='font-family:Arial,sans-serif;max-width:600px;background:#fff;
border:1px solid #e2e2e2;border-radius:10px;overflow:hidden'>
  <div style='background:#e31e24;padding:18px 24px'>
    <h2 style='margin:0;font-size:15px;color:#fff;font-weight:900;letter-spacing:-0.3px'>boom. — New Invoice Submission</h2>
  </div>
  <div style='padding:22px;color:#111'>
    <p style='margin:0 0 14px'>A vendor submitted an invoice through your portal.</p>
    <table style='width:100%;border-collapse:collapse;font-size:13px'>
      {row(True,'Vendor',f"<strong>{vendor_name}</strong>{f' ({vendor_email})' if vendor_email else ''}")}
      {row(False,'Invoice Date',fields.get('invoice_date') or '—')}
      {row(True,'Invoice #',fields.get('invoice_number') or '—')}
      {row(False,'Amount',f"<strong style='color:#e31e24'>{amt_str}</strong>")}
      {row(True,'Description',fields.get('description') or '—')}
      {row(False,'Category',fields.get('category') or '—')}
      {row(True,'Payment Method',fields.get('payment_method') or '—')}
    </table>
    {warn_block}
    <div style='margin-top:20px'>
      <a href='{review_url}' style='background:#e31e24;color:#fff;padding:9px 18px;
border-radius:7px;text-decoration:none;font-weight:600;font-size:13px'>Review in Ledger →</a>
    </div>
  </div>
</div>"""
        resend.Emails.send({"from":"Boom Records <onboarding@resend.dev>",
                            "to":[NOTIFY_EMAIL],
                            "subject":f"New Invoice: {vendor_name} — {amt_str}",
                            "html":html})
    except Exception as e:
        print(f"Email error: {e}")


# ── Main app routes ───────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return render_template("index.html", categories=CATEGORIES,
                           payment_methods=PAYMENT_METHODS, api_configured=bool(ANTHROPIC_KEY),
                           is_admin=is_admin())

@app.route("/parse", methods=["POST"])
@login_required
def parse_invoice():
    if not ANTHROPIC_KEY: return jsonify({"error":"ANTHROPIC_API_KEY not set"}), 400
    if "file" not in request.files: return jsonify({"error":"No file"}), 400
    file = request.files["file"]; file_bytes = file.read()
    ext = Path(file.filename).suffix.lower()
    mime = {".pdf":"application/pdf",".jpg":"image/jpeg",".jpeg":"image/jpeg",
            ".png":"image/png",".webp":"image/webp"}.get(ext,"image/jpeg")
    fields = extract_fields(file_bytes, mime)
    b64 = base64.standard_b64encode(file_bytes).decode()
    preview = f"data:{mime};base64,{b64}" if mime != "application/pdf" else None
    return jsonify({"fields":fields,"preview":preview,"is_pdf":mime=="application/pdf"})

@app.route("/add", methods=["POST"])
@login_required
def add_expense():
    d = request.json
    v = lambda k,df="": (d.get(k,df) or df)
    row = (parse_date(v("invoice_date")), v("payee"), v("description"), v("category"),
           v("artist"), v("song"), v("invoice_number"), parse_amount(v("amount",0)),
           v("payment_method"), parse_date(v("payment_date")), v("in_quickbooks","No"),
           parse_date(v("qb_entry_date")), v("uploaded_to_stem","No"),
           parse_date(v("stem_upload_date")), v("notes"))
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"""INSERT INTO expenses (invoice_date,payee,description,category,
            artist,song,invoice_number,amount,payment_method,payment_date,in_quickbooks,
            qb_entry_date,uploaded_to_stem,stem_upload_date,notes)
            VALUES ({','.join([ph]*15)})""", row)
        new_id = (cur.execute("SELECT lastval()") or cur).fetchone()[0] if kind=="pg" else cur.lastrowid
        conn.commit(); conn.close()
        return jsonify({"ok":True,"id":new_id,"payee":v("payee"),"amount":v("amount")})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/update/<int:eid>", methods=["POST"])
@login_required
def update_entry(eid):
    allowed = {"in_quickbooks","uploaded_to_stem","artist","song","notes",
               "category","payment_method","qb_entry_date","stem_upload_date"}
    updates = {k:v for k,v in request.json.items() if k in allowed}
    if not updates: return jsonify({"error":"No valid fields"}), 400
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        for field, val in updates.items():
            cur.execute(f"UPDATE expenses SET {field}={ph} WHERE id={ph}", (val or None, eid))
        conn.commit(); conn.close()
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/delete/<int:eid>", methods=["POST"])
@login_required
@admin_required
def delete_entry(eid):
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"DELETE FROM expenses WHERE id={ph}", (eid,))
        conn.commit(); conn.close()
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/recent")
@login_required
def recent():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT invoice_date,payee,amount,artist,song,
                              in_quickbooks,uploaded_to_stem
                       FROM expenses ORDER BY id DESC LIMIT 10""")
        rows = cur.fetchall(); conn.close()
        return jsonify([{"date":str(r[0] or ""),"payee":str(r[1] or ""),
                         "amount":r[2],"artist":str(r[3] or ""),"song":str(r[4] or ""),
                         "qb":str(r[5] or ""),"stem":str(r[6] or "")} for r in rows])
    except: return jsonify([])

@app.route("/entries")
@login_required
def entries():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT id,invoice_date,payee,description,category,artist,song,
                              invoice_number,amount,payment_method,payment_date,
                              in_quickbooks,uploaded_to_stem,notes,
                              vendor_submitted,vendor_name
                       FROM expenses ORDER BY invoice_date DESC, id DESC""")
        rows = cur.fetchall(); conn.close()
        return jsonify([{"id":r[0],"invoice_date":str(r[1] or ""),"payee":str(r[2] or ""),
                         "description":str(r[3] or ""),"category":str(r[4] or ""),
                         "artist":str(r[5] or ""),"song":str(r[6] or ""),
                         "invoice_number":str(r[7] or ""),"amount":r[8],
                         "payment_method":str(r[9] or ""),"payment_date":str(r[10] or ""),
                         "in_quickbooks":str(r[11] or ""),"uploaded_to_stem":str(r[12] or ""),
                         "notes":str(r[13] or ""),"vendor_submitted":bool(r[14]),
                         "vendor_name":str(r[15] or "")} for r in rows])
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/ledger")
@login_required
def ledger():
    return render_template("ledger.html", categories=CATEGORIES,
                           payment_methods=PAYMENT_METHODS, is_admin=is_admin())

@app.route("/export")
@login_required
def export_excel():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT invoice_date,payee,description,category,artist,song,
                              invoice_number,amount,payment_method,payment_date,
                              in_quickbooks,qb_entry_date,uploaded_to_stem,stem_upload_date,notes
                       FROM expenses ORDER BY invoice_date ASC, id ASC""")
        rows = cur.fetchall(); conn.close()
    except Exception as e: return jsonify({"error":str(e)}), 500
    wb = _build_excel(rows); buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"BoomRecords_Expenses_{date.today():%Y-%m-%d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _build_excel(rows):
    def fill(c): return PatternFill("solid",start_color=c,end_color=c)
    def bdr():
        s=Side(style="thin",color="FFE2E2E2"); return Border(left=s,right=s,top=s,bottom=s)
    wb=Workbook(); ws=wb.active; ws.title="Expense Tracker"
    ws.sheet_view.showGridLines=False; ws.freeze_panes="A3"
    ws.merge_cells("A1:O1"); ws["A1"]="BOOM RECORDS — EXPENSE & RECOUPMENT TRACKER"
    ws["A1"].font=Font(name="Arial",bold=True,size=13,color="FFFFFFFF")
    ws["A1"].fill=fill("FFE31E24"); ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=28
    hdrs=[("A","Invoice Date",14),("B","Payee / Vendor",22),("C","Description",30),
          ("D","Category",20),("E","Artist / Project",20),("F","Song",20),
          ("G","Invoice #",14),("H","Amount ($)",13),("I","Payment Method",16),
          ("J","Payment Date",14),("K","In QuickBooks?",16),("L","QB Entry Date",14),
          ("M","Uploaded to Stem?",18),("N","Stem Upload Date",16),("O","Notes",30)]
    for col,label,w in hdrs:
        c=ws[f"{col}2"]; c.value=label
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFFFF")
        c.fill=fill("FF333333"); c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=bdr(); ws.column_dimensions[col].width=w
    ws.row_dimensions[2].height=22
    for i,row in enumerate(rows):
        r=i+3; v=list(row)
        qb=str(v[10] or ""); stem=str(v[12] or "")
        rf=fill("FFD9EAD3" if (qb=="Yes" and stem=="Yes") else "FFFFF2CC" if qb=="No" else "FFFCE5CD" if stem=="No" else ("FFF5F5F5" if r%2==0 else "FFFFFFFF"))
        for col,_,_ in hdrs:
            c=ws[f"{col}{r}"]; c.fill=rf
            c.font=Font(name="Arial",size=10); c.border=bdr()
            c.alignment=Alignment(horizontal="left",vertical="center")
        def dc(col,val,fmt=None,align="left"):
            c=ws[f"{col}{r}"]; c.value=val
            if fmt: c.number_format=fmt
            c.alignment=Alignment(horizontal=align,vertical="center")
        dc("A",v[0],"MM/DD/YYYY","center"); dc("B",v[1]); dc("C",v[2]); dc("D",v[3],"","center")
        dc("E",v[4]); dc("F",v[5]); dc("G",v[6],"","center")
        dc("H",v[7],'$#,##0.00;($#,##0.00);"-"',"right")
        dc("I",v[8],"","center"); dc("J",v[9],"MM/DD/YYYY","center"); dc("K",v[10],"","center")
        dc("L",v[11],"MM/DD/YYYY","center"); dc("M",v[12],"","center")
        dc("N",v[13],"MM/DD/YYYY","center"); dc("O",v[14])
    return wb


# ── Vendor submission ─────────────────────────────────────────────────────────

@app.route("/submit", methods=["GET"])
def submit_form():
    return render_template("submit.html")

@app.route("/submit", methods=["POST"])
def submit_invoice():
    vendor_name  = request.form.get("vendor_name","").strip()
    vendor_email = request.form.get("vendor_email","").strip()
    notes        = request.form.get("notes","").strip()
    if not vendor_name:
        return render_template("submit.html", error="Please enter your company or name.")
    if "file" not in request.files or not request.files["file"].filename:
        return render_template("submit.html", error="Please upload your invoice file.")
    file=request.files["file"]; file_bytes=file.read()
    ext=Path(file.filename).suffix.lower()
    mime={".pdf":"application/pdf",".jpg":"image/jpeg",".jpeg":"image/jpeg",
          ".png":"image/png",".webp":"image/webp"}.get(ext,"image/jpeg")
    fields=extract_fields(file_bytes, mime)
    if not fields.get("payee"): fields["payee"]=vendor_name
    unknowns=get_unknowns(fields)
    row=(parse_date(fields.get("invoice_date")),fields.get("payee",""),
         fields.get("description",""),fields.get("category","Other"),
         "","",fields.get("invoice_number",""),parse_amount(fields.get("amount",0)),
         fields.get("payment_method",""),None,"No",None,"No",None,notes,
         True if DATABASE_URL else 1, vendor_name, vendor_email)
    try:
        conn,kind=get_db(); cur=conn.cursor(); ph="%s" if kind=="pg" else "?"
        cur.execute(f"""INSERT INTO expenses (invoice_date,payee,description,category,
            artist,song,invoice_number,amount,payment_method,payment_date,in_quickbooks,
            qb_entry_date,uploaded_to_stem,stem_upload_date,notes,vendor_submitted,
            vendor_name,vendor_email) VALUES ({','.join([ph]*18)})""", row)
        conn.commit(); conn.close()
    except Exception as e:
        return render_template("submit.html", error=f"Submission failed: {e}")
    send_vendor_email(vendor_name, vendor_email, fields, unknowns)
    return render_template("submit_success.html", vendor_name=vendor_name)

@app.route("/status")
def status(): return jsonify({"ok":True})

if __name__ == "__main__":
    init_db()
    port=int(os.environ.get("PORT",5100))
    print(f"\n  Boom Records  →  http://localhost:{port}\n")
    app.run(debug=False, host="0.0.0.0", port=port)
