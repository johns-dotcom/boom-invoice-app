import os, json, base64, io
from datetime import datetime, date
from pathlib import Path
from functools import wraps

from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file, Response
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24))

def _parse_json_list(s):
    """Safely parse a JSON string into a list; return None if empty or invalid."""
    if not s:
        return None
    try:
        parsed = json.loads(s)
        return parsed if isinstance(parsed, list) and parsed else None
    except Exception:
        return None
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024

DATABASE_URL   = os.environ.get("DATABASE_URL", "")
APP_PASSWORD   = os.environ.get("APP_PASSWORD", "")
ANTHROPIC_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")
MODEL          = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")
GMAIL_USER     = os.environ.get("GMAIL_USER", "")        # your Gmail address
GMAIL_APP_PASS = os.environ.get("GMAIL_APP_PASSWORD", "") # Gmail App Password
NOTIFY_EMAIL   = os.environ.get("NOTIFY_EMAIL", "johns@boomrecords.co,jesse@boomrecords.co")
APP_URL        = os.environ.get("APP_URL", "")

# Named admin accounts — each gets their own password and display name.
# Set JOHN_PASSWORD, JESSE_PASSWORD, FELIPE_PASSWORD, SOLI_PASSWORD as env vars in Railway.
# ADMIN_PASSWORD kept as a generic fallback for backward compatibility.
_ADMIN_ACCOUNTS_RAW = [
    (os.environ.get("JOHN_PASSWORD",""),   "John"),
    (os.environ.get("JESSE_PASSWORD",""),  "Jesse"),
    (os.environ.get("FELIPE_PASSWORD",""), "Felipe"),
    (os.environ.get("SOLI_PASSWORD",""),   "Soli"),
    (os.environ.get("ADMIN_PASSWORD",""),  "Admin"),   # generic fallback
]
ADMIN_ACCOUNTS = {pw: name for pw, name in _ADMIN_ACCOUNTS_RAW if pw}

_USER_ACCOUNTS_RAW = [
    (os.environ.get("DANNY_PASSWORD",""), "Danny"),
]
USER_ACCOUNTS = {pw: name for pw, name in _USER_ACCOUNTS_RAW if pw}

# Cobrand is no longer a category — it's a separate yes/no flag on any expense
CATEGORIES = ["Recording","Mixing & Mastering","Music Video","Marketing",
              "Sync/Licensing","Distribution","Legal","Merch","Tour/Live","Other"]
PAYMENT_METHODS = ["ACH","Check","Wire","Credit Card","PayPal","Cash"]

INVOICE_VALIDATE_PROMPT = """Examine this invoice or receipt carefully.
Return ONLY valid JSON — no markdown, no extra text:
{
  "has_invoice_number": true or false,
  "has_amount": true or false,
  "has_date": true or false,
  "has_payee_name": true or false,
  "billed_to_boom": true or false,
  "issues": []
}
Rules:
- Set each boolean to true if that field is clearly present anywhere on the document.
- "billed_to_boom" is true if the invoice is addressed or billed to "Boom.Records", "Boom.Records LLC", or a close variant. False if billed to a different company or person, or if no billing party is shown.
- The "issues" array must ONLY contain entries for fields above that are false. Do NOT flag formatting quirks, empty template fields, or address style.
- If the document is not an invoice or receipt at all, set all to false and add one issue: "This does not appear to be an invoice or receipt."
- Return an empty issues array if all fields are present."""

W9_VALIDATE_PROMPT = """Examine this tax form carefully.
Return ONLY valid JSON — no markdown, no extra text:
{
  "is_w9_or_w8": true or false,
  "has_name": true or false,
  "has_tin_ssn_ein": true or false,
  "has_signature": true or false,
  "has_signed_date": true or false,
  "issues": []
}
Rules:
- "is_w9_or_w8" is true if this is a W-9, W-8BEN, W-8BEN-E, W-8ECI, W-8EXP, W-8IMY, or any variant.
- "has_name" is true if any name (individual or entity) is filled in.
- "has_tin_ssn_ein" is true if ANY of the following: (a) a US SSN, EIN, or ITIN is present in field 5; (b) a foreign tax identifying number (FTIN) is present in field 6a; (c) the "FTIN not legally required" checkbox (field 6b) is checked. Any one of these three satisfies the requirement — treat it as true even when fields 5 and 6a are blank if 6b is checked.
- "has_signature" is true if a handwritten or typed signature appears anywhere on the form.
- "has_signed_date" is true if a date accompanies the signature.
- Always return "issues" as an empty array []. Do not add your own issue commentary."""

EXTRACT_PROMPT = """Extract the following fields from this invoice or receipt.
Return ONLY valid JSON — no markdown, no extra text:
{
  "invoice_date": "MM/DD/YYYY if found, else empty string",
  "payee": "vendor or company name",
  "description": "brief description of what was invoiced (1 sentence max)",
  "category": "best match from: Recording, Mixing & Mastering, Music Video, Marketing, Sync/Licensing, Distribution, Legal, Merch, Tour/Live, Other",
  "artist": "artist, band, or project name referenced anywhere on the invoice (e.g. in a project line, memo, attention field, or description) — empty string if not found",
  "song": "song, album, or project title referenced on the invoice — empty string if not found",
  "invoice_number": "invoice number if present, else empty string",
  "currency": "ISO 4217 currency code (e.g. USD, EUR, GBP, CAD, AUD, MXN, JPY) — default USD if not specified",
  "amount": <number, 2 decimal places, no symbols, 0 if not found>,
  "payment_method": "best match from: ACH, Check, Wire, Credit Card, PayPal, Cash — or empty string"
}"""


# ── DB ────────────────────────────────────────────────────────────────────────

def get_db():
    if DATABASE_URL:
        import psycopg2
        url = DATABASE_URL.replace("postgres://","postgresql://",1)
        return psycopg2.connect(url), "pg"
    else:
        print("WARNING: DATABASE_URL not set — using SQLite. Data will be lost on redeploy!")
        import sqlite3
        conn = sqlite3.connect(str(Path(__file__).parent/"boom.db"))
        conn.row_factory = sqlite3.Row
        return conn, "sqlite"

def init_db():
    conn, kind = get_db()
    cur = conn.cursor()
    if kind == "pg":
        cur.execute("""CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            timestamp TIMESTAMP DEFAULT NOW(),
            user_name TEXT,
            action TEXT,
            entry_id INTEGER,
            entry_payee TEXT,
            field TEXT,
            old_value TEXT,
            new_value TEXT,
            details TEXT)""")
        cur.execute("ALTER TABLE audit_log ADD COLUMN IF NOT EXISTS entry_payee TEXT")
        cur.execute("""CREATE TABLE IF NOT EXISTS expenses (
            id SERIAL PRIMARY KEY, invoice_date DATE, payee TEXT,
            description TEXT, category TEXT, artist TEXT, song TEXT,
            invoice_number TEXT, amount NUMERIC(12,2), payment_method TEXT,
            payment_date DATE, in_quickbooks TEXT DEFAULT 'No',
            qb_entry_date DATE, uploaded_to_stem TEXT DEFAULT 'No',
            stem_upload_date DATE, notes TEXT,
            vendor_submitted BOOLEAN DEFAULT FALSE,
            vendor_name TEXT, vendor_email TEXT,
            w9_filename TEXT, w9_data TEXT,
            invoice_filename TEXT, invoice_data TEXT,
            proof_filename TEXT, proof_data TEXT,
            status TEXT DEFAULT 'approved',
            cobrand BOOLEAN DEFAULT FALSE,
            approved_by TEXT,
            approved_at TIMESTAMP,
            currency TEXT DEFAULT 'USD',
            payment_status TEXT DEFAULT 'Unpaid',
            created_at TIMESTAMP DEFAULT NOW(),
            created_by TEXT)""")
        for col in ["song TEXT","vendor_submitted BOOLEAN DEFAULT FALSE",
                    "vendor_name TEXT","vendor_email TEXT",
                    "w9_filename TEXT","w9_data TEXT",
                    "invoice_filename TEXT","invoice_data TEXT",
                    "proof_filename TEXT","proof_data TEXT",
                    "status TEXT DEFAULT 'approved'",
                    "cobrand BOOLEAN DEFAULT FALSE",
                    "approved_by TEXT",
                    "approved_at TIMESTAMP",
                    "currency TEXT DEFAULT 'USD'",
                    "payment_status TEXT DEFAULT 'Unpaid'",
                    "created_by TEXT",
                    "deleted BOOLEAN DEFAULT FALSE",
                    "vendor_address TEXT",
                    "is_reimbursement BOOLEAN DEFAULT FALSE",
                    "payment_terms TEXT",
                    "artist_breakdown TEXT",
                    "parent_id INTEGER",
                    "boom_rep TEXT"]:
            cur.execute(f"ALTER TABLE expenses ADD COLUMN IF NOT EXISTS {col}")
        cur.execute("UPDATE expenses SET status = 'approved' WHERE status IS NULL")
        cur.execute("UPDATE expenses SET cobrand = FALSE WHERE cobrand IS NULL")
    else:
        cur.execute("""CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT DEFAULT (datetime('now')),
            user_name TEXT,
            action TEXT,
            entry_id INTEGER,
            entry_payee TEXT,
            field TEXT,
            old_value TEXT,
            new_value TEXT,
            details TEXT)""")
        try: cur.execute("ALTER TABLE audit_log ADD COLUMN entry_payee TEXT")
        except: pass
        cur.execute("""CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT, invoice_date TEXT, payee TEXT,
            description TEXT, category TEXT, artist TEXT, song TEXT,
            invoice_number TEXT, amount REAL, payment_method TEXT,
            payment_date TEXT, in_quickbooks TEXT DEFAULT 'No',
            qb_entry_date TEXT, uploaded_to_stem TEXT DEFAULT 'No',
            stem_upload_date TEXT, notes TEXT,
            vendor_submitted INTEGER DEFAULT 0,
            vendor_name TEXT, vendor_email TEXT,
            w9_filename TEXT, w9_data TEXT,
            invoice_filename TEXT, invoice_data TEXT,
            proof_filename TEXT, proof_data TEXT,
            status TEXT DEFAULT 'approved',
            cobrand INTEGER DEFAULT 0,
            approved_by TEXT,
            approved_at TEXT,
            currency TEXT DEFAULT 'USD',
            payment_status TEXT DEFAULT 'Unpaid',
            created_at TEXT DEFAULT (datetime('now')),
            created_by TEXT)""")
        for col in ["song TEXT","vendor_submitted INTEGER DEFAULT 0",
                    "vendor_name TEXT","vendor_email TEXT",
                    "w9_filename TEXT","w9_data TEXT",
                    "invoice_filename TEXT","invoice_data TEXT",
                    "proof_filename TEXT","proof_data TEXT",
                    "status TEXT DEFAULT 'approved'",
                    "cobrand INTEGER DEFAULT 0",
                    "approved_by TEXT",
                    "approved_at TEXT",
                    "currency TEXT DEFAULT 'USD'",
                    "payment_status TEXT DEFAULT 'Unpaid'",
                    "created_by TEXT",
                    "deleted INTEGER DEFAULT 0",
                    "vendor_address TEXT",
                    "is_reimbursement INTEGER DEFAULT 0",
                    "payment_terms TEXT",
                    "artist_breakdown TEXT",
                    "parent_id INTEGER",
                    "boom_rep TEXT"]:
            try: cur.execute(f"ALTER TABLE expenses ADD COLUMN {col}")
            except: pass
        cur.execute("UPDATE expenses SET status = 'approved' WHERE status IS NULL")
        cur.execute("UPDATE expenses SET cobrand = 0 WHERE cobrand IS NULL")
    conn.commit(); conn.close()


# ── Auth ──────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if (APP_PASSWORD or ADMIN_ACCOUNTS) and not session.get("authenticated"):
            return redirect("/login")
        return f(*a, **kw)
    return dec

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if session.get("role") != "admin":
            return jsonify({"error": "Admin access required"}), 403
        return f(*a, **kw)
    return dec

def is_admin():
    return session.get("role") == "admin"

@app.context_processor
def inject_globals():
    if session.get("authenticated"):
        try:
            conn, kind = get_db(); cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM expenses WHERE status = 'pending'")
            count = cur.fetchone()[0]; conn.close()
        except:
            count = 0
        return {"pending_count": count, "current_user": session.get("user_name")}
    return {"pending_count": 0, "current_user": None}

@app.route("/login", methods=["GET","POST"])
def login():
    err = None
    if request.method == "POST":
        pw = request.form.get("password","")
        if pw in ADMIN_ACCOUNTS:
            session["authenticated"] = True
            session["role"] = "admin"
            session["user_name"] = ADMIN_ACCOUNTS[pw]
            return redirect("/ledger")
        elif pw in USER_ACCOUNTS:
            session["authenticated"] = True
            session["role"] = "user"
            session["user_name"] = USER_ACCOUNTS[pw]
            return redirect("/ledger")
        elif APP_PASSWORD and pw == APP_PASSWORD:
            session["authenticated"] = True
            session["role"] = "user"
            session["user_name"] = None
            return redirect("/ledger")
        elif not APP_PASSWORD and not ADMIN_ACCOUNTS:
            session["authenticated"] = True
            session["role"] = "admin"
            session["user_name"] = "Admin"
            return redirect("/ledger")
        else:
            err = "Incorrect password."
    return render_template("login.html", error=err)

@app.route("/logout")
def logout():
    session.clear(); return redirect("/login")

def john_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if session.get("user_name") != "John":
            return redirect("/")
        return f(*a, **kw)
    return dec

def danny_required(f):
    """Danny (or admins) only."""
    @wraps(f)
    def dec(*a, **kw):
        if session.get("user_name") not in ("Danny",) and not is_admin():
            return redirect("/")
        return f(*a, **kw)
    return dec

def history_allowed(f):
    """John, Jesse, and Felipe can view history."""
    @wraps(f)
    def dec(*a, **kw):
        if session.get("user_name") not in ("John", "Jesse", "Felipe"):
            return redirect("/")
        return f(*a, **kw)
    return dec

def log_action(action, entry_id=None, entry_payee=None, field=None, old_value=None, new_value=None, details=None):
    try:
        user = session.get("user_name") or session.get("role") or "unknown"
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"""INSERT INTO audit_log (user_name, action, entry_id, entry_payee, field, old_value, new_value, details)
                        VALUES ({','.join([ph]*8)})""",
                    (user, action, entry_id, entry_payee,
                     field,
                     str(old_value) if old_value is not None else None,
                     str(new_value) if new_value is not None else None,
                     details))
        conn.commit(); conn.close()
    except:
        pass


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_date(s):
    if not s: return None
    for fmt in ("%m/%d/%Y","%Y-%m-%d","%m-%d-%Y","%d/%m/%Y"):
        try: return datetime.strptime(s.strip(), fmt).date()
        except: pass
    return None

def parse_amount(v):
    try: return float(str(v).replace("$","").replace(",","").strip()) if v else None
    except: return None

def ext_mime(filename):
    ext = Path(filename).suffix.lower() if filename else ""
    return {".pdf":"application/pdf",".jpg":"image/jpeg",".jpeg":"image/jpeg",
            ".png":"image/png",".webp":"image/webp"}.get(ext,"application/octet-stream")

def extract_fields(file_bytes, mime):
    if not ANTHROPIC_KEY: return {}
    b64 = base64.standard_b64encode(file_bytes).decode()
    content = ([{"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64}},
                {"type":"text","text":EXTRACT_PROMPT}]
               if mime == "application/pdf" else
               [{"type":"image","source":{"type":"base64","media_type":mime,"data":b64}},
                {"type":"text","text":EXTRACT_PROMPT}])
    try:
        resp = anthropic.Anthropic(api_key=ANTHROPIC_KEY).messages.create(
            model=MODEL, max_tokens=512, messages=[{"role":"user","content":content}])
        raw = resp.content[0].text.strip()
        if raw.startswith("```"): raw = raw.split("```")[1]; raw = raw[4:] if raw.startswith("json") else raw
        return json.loads(raw.strip())
    except Exception as e:
        print(f"Extraction error: {e}"); return {}

def get_unknowns(fields):
    u = []
    if not fields.get("invoice_date"): u.append("Invoice date missing")
    if not fields.get("amount") or float(fields.get("amount",0))==0: u.append("Amount not found or zero")
    if not fields.get("invoice_number"): u.append("Invoice number missing")
    if fields.get("category") in ("Other","",None): u.append("Category could not be determined")
    if not fields.get("description"): u.append("Description missing")
    return u

def serve_file(data_b64, filename, inline=True):
    if not data_b64: return "File not found", 404
    file_bytes = base64.b64decode(data_b64)
    mime = ext_mime(filename)
    disposition = "inline" if inline and mime in ("application/pdf","image/jpeg","image/png","image/webp") else "attachment"
    resp = Response(file_bytes, mimetype=mime)
    resp.headers["Content-Disposition"] = f'{disposition}; filename="{filename}"'
    return resp

def fmt_qbo_date(d):
    if d is None: return datetime.now().strftime("%Y%m%d") + "120000"
    if hasattr(d, 'strftime'): return d.strftime("%Y%m%d") + "120000"
    s = str(d).replace("-","")[:8]
    return s + "120000" if len(s) >= 8 else datetime.now().strftime("%Y%m%d") + "120000"


# ── Email ─────────────────────────────────────────────────────────────────────

def send_vendor_email(vendor_name, vendor_email, fields, unknowns, w9_filename=None, is_reimbursement=False):
    client_id     = os.environ.get("GMAIL_CLIENT_ID", "")
    client_secret = os.environ.get("GMAIL_CLIENT_SECRET", "")
    refresh_token = os.environ.get("GMAIL_REFRESH_TOKEN", "")
    sender        = os.environ.get("GMAIL_USER", "")
    if not all([client_id, client_secret, refresh_token, sender]):
        print("Email not configured — set GMAIL_CLIENT_ID, GMAIL_CLIENT_SECRET, GMAIL_REFRESH_TOKEN, GMAIL_USER")
        return
    try:
        import base64, urllib.request, urllib.parse, json
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        # Exchange refresh token for a fresh access token
        token_resp = urllib.request.urlopen(urllib.request.Request(
            "https://oauth2.googleapis.com/token",
            data=urllib.parse.urlencode({
                "client_id":     client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token,
                "grant_type":    "refresh_token",
            }).encode(),
            method="POST"
        ))
        access_token = json.loads(token_resp.read())["access_token"]

        review_url = f"{APP_URL}/approvals" if APP_URL else "https://boomap.com/approvals"
        amt = fields.get("amount", 0)
        amt_str = f"${float(amt):,.2f}" if amt else "Unknown"

        warn = ("".join(f"<li style='color:#d97706'>⚠ {u}</li>" for u in unknowns) if unknowns else "")
        warn_block = (f"<div style='background:#fef3c7;border:1px solid #fcd34d;border-radius:8px;"
                      f"padding:12px 16px;margin:14px 0'><strong style='color:#92400e'>Needs review:</strong>"
                      f"<ul style='margin:6px 0 0 16px;color:#92400e'>{warn}</ul></div>") if warn else ""
        w9_block = (f"<div style='background:#f0fdf4;border:1px solid #86efac;border-radius:8px;"
                    f"padding:10px 14px;margin:14px 0;font-size:13px;color:#166534'>"
                    f"W9/W8 submitted: <strong>{w9_filename}</strong></div>") if w9_filename else ""

        def row(bg, label, val):
            bg_s = "background:#f9f9f9;" if bg else ""
            return (f"<tr><td style='padding:7px 12px;{bg_s}color:#666;width:150px'>{label}</td>"
                    f"<td style='padding:7px 12px;{bg_s}'>{val}</td></tr>")

        doc_type = "Reimbursement" if is_reimbursement else "Invoice"
        html = f"""<div style='font-family:Arial,sans-serif;max-width:600px;background:#fff;
border:1px solid #e2e2e2;border-radius:10px;overflow:hidden'>
  <div style='background:#e31e24;padding:18px 24px'>
    <h2 style='margin:0;font-size:15px;color:#fff;font-weight:900'>boom. — New {doc_type} Pending Approval</h2>
  </div>
  <div style='padding:22px;color:#111'>
    <p style='margin:0 0 14px'>A vendor submitted an invoice — it's waiting in your approval queue.</p>
    <table style='width:100%;border-collapse:collapse;font-size:13px'>
      {row(True,'Vendor',f"<strong>{vendor_name}</strong> ({vendor_email})")}
      {row(False,'Artist / Project',f"<strong>{fields.get('artist') or '—'}</strong>")}
      {row(True,'Song',fields.get('song') or '—')}
      {row(False,'Invoice Date',fields.get('invoice_date') or '—')}
      {row(False,'Invoice #',fields.get('invoice_number') or '—')}
      {row(True,'Amount',f"<strong style='color:#e31e24'>{amt_str}</strong>")}
      {row(False,'Description',fields.get('description') or '—')}
      {row(True,'Category',fields.get('category') or '—')}
      {row(False,'Boom Rep',fields.get('boom_rep') or '—')}
    </table>
    {w9_block}{warn_block}
    <div style='margin-top:20px'>
      <a href='{review_url}' style='background:#e31e24;color:#fff;padding:9px 18px;
border-radius:7px;text-decoration:none;font-weight:600;font-size:13px'>Review &amp; Approve</a>
    </div>
  </div>
</div>"""

        recipients = [e.strip() for e in NOTIFY_EMAIL.split(",") if e.strip()]
        msg = MIMEMultipart("alternative")
        song_str = fields.get('song','').strip()
        subject_artist = fields.get('artist','').strip()
        subject_detail = f"{subject_artist} — {song_str}" if subject_artist and song_str else (subject_artist or song_str or vendor_name)
        msg["Subject"] = f"New {doc_type}: {vendor_name} · {subject_detail} · {amt_str}"
        msg["From"]    = f"Boom.Records <{sender}>"
        msg["To"]      = ", ".join(recipients)
        msg.attach(MIMEText(html, "html"))

        # Send via Gmail API (no SMTP, no extra libraries)
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        urllib.request.urlopen(urllib.request.Request(
            "https://gmail.googleapis.com/gmail/v1/users/me/messages/send",
            data=json.dumps({"raw": raw}).encode(),
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type":  "application/json",
            },
            method="POST"
        ))
        print(f"Vendor email sent to {recipients}")

    except Exception as e:
        print(f"Email error: {e}")


def boom_rep_email(name):
    """Return the Boom Rep's email address by name."""
    if not name:
        return None
    special = {"john": "johns@boomrecords.co"}
    key = name.lower().strip()
    return special.get(key, f"{key}@boomrecords.co")


def send_status_email(vendor_name, vendor_email, status, invoice_info, boom_rep=None, reason=None):
    """Send an approval or rejection email directly to the vendor."""
    client_id     = os.environ.get("GMAIL_CLIENT_ID", "")
    client_secret = os.environ.get("GMAIL_CLIENT_SECRET", "")
    refresh_token = os.environ.get("GMAIL_REFRESH_TOKEN", "")
    sender        = os.environ.get("GMAIL_USER", "")
    if not all([client_id, client_secret, refresh_token, sender]):
        print("Email not configured — vendor status email not sent")
        return
    if not vendor_email:
        print("No vendor email — vendor status email not sent")
        return
    try:
        import base64, urllib.request, urllib.parse, json
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        token_resp = urllib.request.urlopen(urllib.request.Request(
            "https://oauth2.googleapis.com/token",
            data=urllib.parse.urlencode({
                "client_id":     client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token,
                "grant_type":    "refresh_token",
            }).encode(), method="POST"
        ))
        access_token = json.loads(token_resp.read())["access_token"]

        rep_name  = boom_rep or "the Boom.Records team"
        rep_email = boom_rep_email(boom_rep) if boom_rep else None
        contact_line = (f'<strong>{rep_name}</strong> at '
                        f'<a href="mailto:{rep_email}" style="color:#e31e24">{rep_email}</a>'
                        if rep_email else f'<strong>{rep_name}</strong>')

        amt     = invoice_info.get("amount")
        amt_str = f"${float(amt):,.2f}" if amt else ""
        inv_num = invoice_info.get("invoice_number", "")
        artist  = invoice_info.get("artist", "")

        def row(bg, label, val):
            s = "background:#f9f9f9;" if bg else ""
            return (f"<tr><td style='padding:7px 12px;{s}color:#666;width:130px;font-size:13px'>{label}</td>"
                    f"<td style='padding:7px 12px;{s}font-size:13px'>{val}</td></tr>")

        if status == "approved":
            subject      = f"Invoice #{inv_num} Approved — Boom.Records" if inv_num else "Invoice Approved — Boom.Records"
            accent_color = "#16a34a"
            header_txt   = "Your invoice has been approved ✓"
            body_html    = f"""
<p style="margin:0 0 14px">Hi <strong>{vendor_name}</strong>,</p>
<p style="margin:0 0 14px">Your invoice has been reviewed and approved. Payment will be processed according to the agreed terms.</p>"""
        else:
            subject      = f"Invoice #{inv_num} — Follow-Up Needed" if inv_num else "Invoice Submission — Follow-Up Needed"
            accent_color = "#e31e24"
            header_txt   = "Follow-up needed on your invoice"
            reason_block = (f"<div style='background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;"
                            f"padding:12px 16px;margin:14px 0;font-size:13px;color:#b91c1c'>"
                            f"<strong>Reason:</strong> {reason}</div>") if reason else ""
            body_html    = f"""
<p style="margin:0 0 14px">Hi <strong>{vendor_name}</strong>,</p>
<p style="margin:0 0 14px">Thank you for submitting your invoice. Unfortunately we weren't able to process it at this time.</p>
{reason_block}
<p style="margin:0 0 14px">Please reach out to your Boom Rep for next steps.</p>"""

        detail_rows = ""
        if artist:  detail_rows += row(True,  "Artist / Project", f"<strong>{artist}</strong>")
        if inv_num: detail_rows += row(False, "Invoice #", inv_num)
        if amt_str: detail_rows += row(True,  "Amount", f"<strong style='color:{accent_color}'>{amt_str}</strong>")
        details_table = (f"<table style='width:100%;border-collapse:collapse;margin:14px 0'>"
                         f"{detail_rows}</table>") if detail_rows else ""

        html = f"""<div style='font-family:Arial,sans-serif;max-width:600px;background:#fff;
border:1px solid #e2e2e2;border-radius:10px;overflow:hidden'>
  <div style='background:{accent_color};padding:18px 24px'>
    <h2 style='margin:0;font-size:15px;color:#fff;font-weight:900'>boom. — {header_txt}</h2>
  </div>
  <div style='padding:22px;color:#111'>
    {body_html}
    {details_table}
    <p style='margin:14px 0 0;font-size:13px;color:#555'>
      Questions? Reach out to your Boom Rep: {contact_line}</p>
    <p style='margin:8px 0 0;font-size:12px;color:#aaa'>Boom.Records LLC</p>
  </div>
</div>"""

        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"Boom.Records <{sender}>"
        msg["To"]      = vendor_email
        if status == "approved" and rep_email:
            msg["Cc"] = rep_email
        msg.attach(MIMEText(html, "html"))

        # Build recipient list (vendor + boom rep CC on approvals)
        recipients = [vendor_email]
        if status == "approved" and rep_email:
            recipients.append(rep_email)

        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        urllib.request.urlopen(urllib.request.Request(
            "https://gmail.googleapis.com/gmail/v1/users/me/messages/send",
            data=json.dumps({"raw": raw, "deliveryReceipt": False}).encode(),
            headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"},
            method="POST"
        ))
        print(f"Vendor status email ({status}) sent to {recipients}")
    except Exception as e:
        print(f"Vendor status email error: {e}")


# ── Health check ──────────────────────────────────────────────────────────────

@app.route("/health")
def health():
    db_type = "postgresql" if DATABASE_URL else "sqlite (WARNING: no DATABASE_URL set!)"
    try:
        conn, kind = get_db()
        conn.cursor().execute("SELECT 1")
        conn.close()
        status = "connected"
    except Exception as e:
        status = f"ERROR: {e}"
    return jsonify({"database": db_type, "status": status, "database_url_set": bool(DATABASE_URL)})

# ── Main app routes ───────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return redirect(url_for("ledger"))

@app.route("/add")
@login_required
def invoice_form():
    return render_template("index.html", categories=CATEGORIES,
                           payment_methods=PAYMENT_METHODS,
                           api_configured=bool(ANTHROPIC_KEY),
                           is_admin=is_admin())

@app.route("/parse", methods=["POST"])
@login_required
def parse_invoice():
    if not ANTHROPIC_KEY: return jsonify({"error":"ANTHROPIC_API_KEY not set"}), 400
    if "file" not in request.files: return jsonify({"error":"No file"}), 400
    file = request.files["file"]; file_bytes = file.read()
    fname = file.filename
    mime = ext_mime(fname)
    fields = extract_fields(file_bytes, mime)
    b64 = base64.standard_b64encode(file_bytes).decode()
    preview = f"data:{mime};base64,{b64}" if mime != "application/pdf" else None
    return jsonify({"fields":fields,"preview":preview,"is_pdf":mime=="application/pdf",
                    "file_b64":b64,"file_mime":mime,"file_name":fname})

@app.route("/add", methods=["POST"])
@login_required
def add_expense():
    d = request.json
    v = lambda k,df="": (d.get(k,df) or df)
    cobrand = bool(d.get("cobrand"))
    currency = (v("currency") or "USD").upper().strip()[:3]
    added_by = session.get("user_name") or session.get("role") or "unknown"
    contact_email = v("contact_email") or None
    row = (parse_date(v("invoice_date")), v("payee"), v("description"), v("category"),
           v("artist"), v("song"), v("invoice_number"), parse_amount(v("amount",0)),
           v("payment_method"), parse_date(v("payment_date")), v("in_quickbooks","No"),
           parse_date(v("qb_entry_date")), v("uploaded_to_stem","No"),
           parse_date(v("stem_upload_date")), v("notes"),
           v("invoice_filename") or None, v("invoice_b64") or None, cobrand,
           v("w9_filename") or None, v("w9_b64") or None, currency, added_by, contact_email)
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        # Duplicate check — only if invoice_number is provided and force not set
        if v("invoice_number") and not d.get("force"):
            cur.execute(f"""SELECT id FROM expenses WHERE invoice_number={ph} AND payee={ph}
                            AND deleted IS NOT TRUE LIMIT 1""",
                        (v("invoice_number"), v("payee")))
            existing = cur.fetchone()
            if existing:
                conn.close()
                return jsonify({"duplicate": True,
                                "message": f"Invoice #{v('invoice_number')} from {v('payee')} is already in the ledger (entry #{existing[0]})."}), 409
        cur.execute(f"""INSERT INTO expenses (invoice_date,payee,description,category,
            artist,song,invoice_number,amount,payment_method,payment_date,in_quickbooks,
            qb_entry_date,uploaded_to_stem,stem_upload_date,notes,invoice_filename,invoice_data,cobrand,
            w9_filename,w9_data,currency,created_by,vendor_email)
            VALUES ({','.join([ph]*23)})""", row)
        new_id = (cur.execute("SELECT lastval()") or cur).fetchone()[0] if kind=="pg" else cur.lastrowid
        conn.commit(); conn.close()
        log_action("invoice_added", new_id, v("payee"),
                   details=f"Invoice #{v('invoice_number')} | {currency} {v('amount')} | {v('category')}")
        return jsonify({"ok":True,"id":new_id,"payee":v("payee"),"amount":v("amount")})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/update/<int:eid>", methods=["POST"])
@login_required
def update_entry(eid):
    allowed = {"in_quickbooks","uploaded_to_stem","artist","song","notes",
               "category","payment_method","payment_date","qb_entry_date","stem_upload_date","cobrand","currency","payment_status","vendor_email","payment_terms","invoice_number","boom_rep",
               "payee","amount","description","invoice_date"}
    updates = {k:v for k,v in request.json.items() if k in allowed}
    if not updates: return jsonify({"error":"No valid fields"}), 400
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        # Fetch old values + payee for audit log
        fields_list = list(updates.keys())
        cur.execute(f"SELECT payee,{','.join(fields_list)} FROM expenses WHERE id={ph}", (eid,))
        old_row = cur.fetchone()
        payee_val = old_row[0] if old_row else ""
        old_vals = {fields_list[i]: old_row[i+1] for i in range(len(fields_list))} if old_row else {}
        for field, val in updates.items():
            if field == "cobrand":
                val = bool(val)
            elif field == "amount":
                try: val = float(str(val).replace(",","").replace("$","").strip()) if val not in (None,"") else None
                except: val = None
            cur.execute(f"UPDATE expenses SET {field}={ph} WHERE id={ph}", (val if val != "" else None, eid))
            log_action("field_updated", eid, payee_val, field=field,
                       old_value=old_vals.get(field), new_value=val)
        conn.commit(); conn.close()
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/add-invoice/<int:eid>", methods=["POST"])
@login_required
def add_invoice(eid):
    if "file" not in request.files or not request.files["file"].filename:
        return jsonify({"error":"No file"}), 400
    f = request.files["file"]
    fname = f.filename
    data = base64.b64encode(f.read()).decode()
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"SELECT payee FROM expenses WHERE id={ph}", (eid,))
        payee_row = cur.fetchone()
        cur.execute(f"UPDATE expenses SET invoice_filename={ph}, invoice_data={ph} WHERE id={ph}",
                    (fname, data, eid))
        conn.commit(); conn.close()
        log_action("invoice_file_uploaded", eid, payee_row[0] if payee_row else None, details=fname)
        return jsonify({"ok":True,"filename":fname})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/add-w9/<int:eid>", methods=["POST"])
@login_required
def add_w9(eid):
    if "file" not in request.files or not request.files["file"].filename:
        return jsonify({"error":"No file"}), 400
    f = request.files["file"]
    fname = f.filename
    data = base64.b64encode(f.read()).decode()
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"SELECT payee FROM expenses WHERE id={ph}", (eid,))
        payee_row = cur.fetchone()
        cur.execute(f"UPDATE expenses SET w9_filename={ph}, w9_data={ph} WHERE id={ph}",
                    (fname, data, eid))
        conn.commit(); conn.close()
        log_action("w9_file_uploaded", eid, payee_row[0] if payee_row else None, details=fname)
        return jsonify({"ok":True,"filename":fname})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/w9-only", methods=["POST"])
@login_required
def w9_only():
    """Save a standalone W9 — creates a minimal approved expense record."""
    payee = request.form.get("payee","").strip()
    if not payee:
        return jsonify({"error":"Payee name required"}), 400
    w9_f = request.files.get("file")
    if not w9_f or not w9_f.filename:
        return jsonify({"error":"W9 file required"}), 400
    w9_fname = w9_f.filename
    w9_data  = base64.b64encode(w9_f.read()).decode()
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"""INSERT INTO expenses (payee, w9_filename, w9_data, status)
                        VALUES ({ph},{ph},{ph},{ph})""", (payee, w9_fname, w9_data, "approved"))
        new_id = (cur.execute("SELECT lastval()") or cur).fetchone()[0] if kind=="pg" else cur.lastrowid
        conn.commit(); conn.close()
        log_action("w9_submitted", new_id, payee, details=w9_fname)
        return jsonify({"ok":True,"id":new_id,"payee":payee})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/lookup-invoice", methods=["GET"])
@login_required
def lookup_invoice():
    """Look up an expense by invoice number — used for proof-of-payment matching."""
    number = request.args.get("number","").strip()
    if not number:
        return jsonify({"found":False})
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"SELECT id, payee, amount FROM expenses WHERE invoice_number={ph} LIMIT 1", (number,))
        row = cur.fetchone(); conn.close()
        if row:
            return jsonify({"found":True,"id":row[0],"payee":str(row[1] or ""),"amount":str(row[2] or "")})
        return jsonify({"found":False})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/add-proof-standalone", methods=["POST"])
@login_required
def add_proof_standalone():
    """Attach proof of payment to an existing invoice (by matched_id) or create a new record."""
    payee     = request.form.get("payee","").strip()
    inv_num   = request.form.get("invoice_number","").strip()
    matched_id = request.form.get("matched_id","").strip()
    if not payee:
        return jsonify({"error":"Payee name required"}), 400
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error":"Proof file required"}), 400
    fname = f.filename
    data  = base64.b64encode(f.read()).decode()
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        if matched_id:
            # Attach to existing record
            cur.execute(f"UPDATE expenses SET proof_filename={ph}, proof_data={ph} WHERE id={ph}",
                        (fname, data, int(matched_id)))
            conn.commit(); conn.close()
            return jsonify({"ok":True,"matched":True,"payee":payee})
        else:
            # Create a new record
            cur.execute(f"""INSERT INTO expenses (payee, invoice_number, proof_filename, proof_data, status)
                            VALUES ({ph},{ph},{ph},{ph},{ph})""",
                        (payee, inv_num or None, fname, data, "approved"))
            new_id = (cur.execute("SELECT lastval()") or cur).fetchone()[0] if kind=="pg" else cur.lastrowid
            conn.commit(); conn.close()
            return jsonify({"ok":True,"matched":False,"id":new_id,"payee":payee})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/add-proof/<int:eid>", methods=["POST"])
@login_required
def add_proof(eid):
    if "file" not in request.files or not request.files["file"].filename:
        return jsonify({"error":"No file"}), 400
    f = request.files["file"]
    fname = f.filename
    data = base64.b64encode(f.read()).decode()
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"UPDATE expenses SET proof_filename={ph}, proof_data={ph} WHERE id={ph}",
                    (fname, data, eid))
        conn.commit(); conn.close()
        return jsonify({"ok":True,"filename":fname})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/remove-invoice/<int:eid>", methods=["POST"])
@login_required
def remove_invoice(eid):
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"SELECT payee, invoice_filename FROM expenses WHERE id={ph}", (eid,))
        row = cur.fetchone()
        cur.execute(f"UPDATE expenses SET invoice_filename=NULL, invoice_data=NULL WHERE id={ph}", (eid,))
        conn.commit(); conn.close()
        log_action("invoice_file_removed", eid, row[0] if row else None, details=row[1] if row else None)
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/remove-w9/<int:eid>", methods=["POST"])
@login_required
def remove_w9(eid):
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"SELECT payee, w9_filename FROM expenses WHERE id={ph}", (eid,))
        row = cur.fetchone()
        cur.execute(f"UPDATE expenses SET w9_filename=NULL, w9_data=NULL WHERE id={ph}", (eid,))
        conn.commit(); conn.close()
        log_action("w9_file_removed", eid, row[0] if row else None, details=row[1] if row else None)
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/remove-proof/<int:eid>", methods=["POST"])
@login_required
def remove_proof(eid):
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"SELECT payee, proof_filename FROM expenses WHERE id={ph}", (eid,))
        row = cur.fetchone()
        cur.execute(f"UPDATE expenses SET proof_filename=NULL, proof_data=NULL WHERE id={ph}", (eid,))
        conn.commit(); conn.close()
        log_action("proof_file_removed", eid, row[0] if row else None, details=row[1] if row else None)
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/delete/<int:eid>", methods=["POST"])
@login_required
@admin_required
def delete_entry(eid):
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"SELECT payee, invoice_number FROM expenses WHERE id={ph}", (eid,))
        row = cur.fetchone()
        cur.execute(f"UPDATE expenses SET deleted=TRUE WHERE id={ph}", (eid,))
        conn.commit(); conn.close()
        log_action("invoice_deleted", eid, row[0] if row else None,
                   details=f"Invoice #{row[1]}" if row and row[1] else None)
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/restore/<int:eid>", methods=["POST"])
@login_required
@admin_required
def restore_entry(eid):
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"UPDATE expenses SET deleted=FALSE WHERE id={ph}", (eid,))
        conn.commit(); conn.close()
        log_action("invoice_restored", eid)
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}), 500


# ── Approval queue ────────────────────────────────────────────────────────────

@app.route("/approvals")
@login_required
def approvals_page():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT id, created_at, vendor_name, vendor_email,
                              invoice_date, payee, description, category,
                              invoice_number, amount, notes,
                              invoice_filename, w9_filename, artist, song, cobrand,
                              is_reimbursement, artist_breakdown, boom_rep
                       FROM expenses WHERE status = 'pending'
                       ORDER BY created_at ASC""")
        rows = cur.fetchall(); conn.close()
        items = [{"id":r[0],"created_at":str(r[1] or ""),"vendor_name":str(r[2] or ""),
                  "vendor_email":str(r[3] or ""),"invoice_date":str(r[4] or ""),
                  "payee":str(r[5] or ""),"description":str(r[6] or ""),
                  "category":str(r[7] or ""),"invoice_number":str(r[8] or ""),
                  "amount":r[9],"notes":str(r[10] or ""),
                  "invoice_filename":str(r[11] or ""),"w9_filename":str(r[12] or ""),
                  "artist":str(r[13] or ""),"song":str(r[14] or ""),
                  "cobrand":bool(r[15]),"is_reimbursement":bool(r[16]),
                  "artist_breakdown": _parse_json_list(r[17]),
                  "boom_rep": str(r[18] or "")} for r in rows]
    except Exception as e:
        items = []
    return render_template("approvals.html", items=items, is_admin=is_admin())

@app.route("/approve/<int:eid>", methods=["POST"])
@login_required
@admin_required
def approve_entry(eid):
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        approver = session.get("user_name") or "Admin"
        now = datetime.now()

        # Fetch full row so we can clone it for split entries
        cur.execute(f"""SELECT payee, invoice_number, amount, artist_breakdown,
                               invoice_date, description, category,
                               payment_method, notes, vendor_name, vendor_email,
                               vendor_address, w9_filename, w9_data,
                               invoice_filename, invoice_data,
                               proof_filename, proof_data,
                               cobrand, is_reimbursement, currency, payment_terms, created_at,
                               boom_rep
                        FROM expenses WHERE id={ph}""", (eid,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Entry not found"}), 404

        (payee, inv_num, total_amount, breakdown_str,
         inv_date, desc, category,
         pay_method, notes, vendor_name, vendor_email,
         vendor_addr, w9_fname, w9_data_val,
         inv_fname, inv_data_val,
         proof_fname, proof_data_val,
         cobrand, is_reimb, currency, pay_terms, created_at_val,
         boom_rep) = row

        breakdown = _parse_json_list(breakdown_str)

        if breakdown and len(breakdown) > 1:
            # ── Multi-artist split ───────────────────────────────────────────
            n = len(breakdown)
            raw_amounts = [e.get("amount") for e in breakdown]
            # "Has amounts" = at least one row has a non-zero, non-null amount
            has_any_amount = any(
                a is not None and str(a).strip() not in ("", "0", "0.0", "0.00")
                for a in raw_amounts
            )
            if has_any_amount:
                split_amounts = []
                for a in raw_amounts:
                    try:
                        split_amounts.append(
                            round(float(str(a).replace(",","").strip()), 2)
                            if a not in (None, "") else 0.0
                        )
                    except:
                        split_amounts.append(0.0)
            else:
                # Even split — give any rounding remainder to the first artist
                base = float(total_amount or 0)
                per  = round(base / n, 2)
                remainder = round(base - per * n, 2)
                split_amounts = [round(per + remainder, 2)] + [per] * (n - 1)

            # Update original row → first artist's data, clear breakdown
            cur.execute(f"""UPDATE expenses
                             SET status='approved', approved_by={ph}, approved_at={ph},
                                 artist={ph}, song={ph}, amount={ph}, artist_breakdown=NULL
                            WHERE id={ph}""",
                        (approver, now,
                         breakdown[0].get("artist",""), breakdown[0].get("song",""),
                         split_amounts[0], eid))

            # Insert a cloned row for each remaining artist
            for i, entry in enumerate(breakdown[1:], 1):
                cur.execute(f"""INSERT INTO expenses (
                    invoice_date, payee, description, category, invoice_number,
                    payment_method, notes, vendor_name, vendor_email, vendor_address,
                    w9_filename, w9_data, invoice_filename, invoice_data,
                    proof_filename, proof_data,
                    cobrand, is_reimbursement, currency, payment_terms, created_at,
                    artist, song, amount,
                    status, approved_by, approved_at, parent_id
                ) VALUES ({','.join([ph]*28)})""",
                (inv_date, payee, desc, category, inv_num,
                 pay_method, notes, vendor_name, vendor_email, vendor_addr,
                 w9_fname, w9_data_val, inv_fname, inv_data_val,
                 proof_fname, proof_data_val,
                 cobrand, is_reimb, currency, pay_terms, created_at_val,
                 entry.get("artist",""), entry.get("song",""), split_amounts[i],
                 "approved", approver, now, eid))
        else:
            # ── Single artist — standard approval ────────────────────────────
            cur.execute(f"""UPDATE expenses SET status='approved', approved_by={ph}, approved_at={ph}
                            WHERE id={ph}""", (approver, now, eid))

        conn.commit(); conn.close()
        detail_parts = []
        if inv_num: detail_parts.append(f"Invoice #{inv_num}")
        if total_amount: detail_parts.append(f"${total_amount}")
        log_action("invoice_approved", eid, payee,
                   details=" | ".join(detail_parts) if detail_parts else None)
        # Email vendor
        first_artist = breakdown[0].get("artist","") if breakdown else payee
        send_status_email(
            vendor_name or payee, vendor_email, "approved",
            {"invoice_number": inv_num, "amount": total_amount, "artist": first_artist},
            boom_rep=boom_rep
        )
        return jsonify({"ok":True, "approved_by": approver, "approved_at": now.strftime("%Y-%m-%d %H:%M")})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/resplit/<int:eid>", methods=["POST"])
@login_required
@admin_required
def resplit_entry(eid):
    """Retroactively split an already-approved single entry into multiple artist rows."""
    try:
        data = request.json or {}
        breakdown = data.get("breakdown", [])
        if not breakdown or len(breakdown) < 2:
            return jsonify({"error": "Need at least 2 artists"}), 400

        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"

        # Fetch the full existing row
        cur.execute(f"""SELECT payee, invoice_number, amount,
                               invoice_date, description, category,
                               payment_method, notes, vendor_name, vendor_email,
                               vendor_address, w9_filename, w9_data,
                               invoice_filename, invoice_data,
                               proof_filename, proof_data,
                               cobrand, is_reimbursement, currency, payment_terms,
                               created_at, status, approved_by, approved_at
                        FROM expenses WHERE id={ph}""", (eid,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Entry not found"}), 404

        (payee, inv_num, total_amount,
         inv_date, desc, category,
         pay_method, notes, vendor_name, vendor_email,
         vendor_addr, w9_fname, w9_data_val,
         inv_fname, inv_data_val,
         proof_fname, proof_data_val,
         cobrand, is_reimb, currency, pay_terms,
         created_at_val, status, approved_by, approved_at) = row

        # Compute amounts — use provided amounts if any, else split evenly
        raw_amounts = [e.get("amount") for e in breakdown]
        has_any = any(a not in (None, "", 0, 0.0) for a in raw_amounts)
        if has_any:
            split_amounts = []
            for a in raw_amounts:
                try:
                    split_amounts.append(round(float(str(a).replace(",","").strip()), 2) if a not in (None,"") else 0.0)
                except:
                    split_amounts.append(0.0)
        else:
            n = len(breakdown)
            base = float(total_amount or 0)
            per = round(base / n, 2)
            remainder = round(base - per * n, 2)
            split_amounts = [round(per + remainder, 2)] + [per] * (n - 1)

        # Delete any existing child rows (in case this is a re-do)
        cur.execute(f"DELETE FROM expenses WHERE parent_id={ph}", (eid,))

        # Update original row to first artist
        cur.execute(f"""UPDATE expenses
                         SET artist={ph}, song={ph}, amount={ph}, artist_breakdown=NULL
                        WHERE id={ph}""",
                    (breakdown[0].get("artist",""), breakdown[0].get("song",""),
                     split_amounts[0], eid))

        # Insert cloned rows for remaining artists
        for i, entry in enumerate(breakdown[1:], 1):
            cur.execute(f"""INSERT INTO expenses (
                invoice_date, payee, description, category, invoice_number,
                payment_method, notes, vendor_name, vendor_email, vendor_address,
                w9_filename, w9_data, invoice_filename, invoice_data,
                proof_filename, proof_data,
                cobrand, is_reimbursement, currency, payment_terms, created_at,
                artist, song, amount,
                status, approved_by, approved_at, parent_id
            ) VALUES ({','.join([ph]*28)})""",
            (inv_date, payee, desc, category, inv_num,
             pay_method, notes, vendor_name, vendor_email, vendor_addr,
             w9_fname, w9_data_val, inv_fname, inv_data_val,
             proof_fname, proof_data_val,
             cobrand, is_reimb, currency, pay_terms, created_at_val,
             entry.get("artist",""), entry.get("song",""), split_amounts[i],
             status, approved_by, approved_at, eid))

        conn.commit(); conn.close()
        log_action("entry_resplit", eid, payee, details=f"{len(breakdown)} artists")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/approve-bulk", methods=["POST"])
@login_required
@admin_required
def approve_bulk():
    try:
        data = request.json or {}
        ids = data.get("ids", [])
        if not ids:
            return jsonify({"ok": False, "error": "No IDs provided"}), 400

        approved = []
        failed = []

        for eid in ids:
            try:
                conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
                approver = session.get("user_name") or "Admin"
                now = datetime.now()

                # Fetch full row
                cur.execute(f"""SELECT payee, invoice_number, amount, artist_breakdown,
                                       invoice_date, description, category,
                                       payment_method, notes, vendor_name, vendor_email,
                                       vendor_address, w9_filename, w9_data,
                                       invoice_filename, invoice_data,
                                       proof_filename, proof_data,
                                       cobrand, is_reimbursement, currency, payment_terms, created_at
                                FROM expenses WHERE id={ph}""", (eid,))
                row = cur.fetchone()
                if not row:
                    conn.close()
                    failed.append(eid)
                    continue

                (payee, inv_num, total_amount, breakdown_str,
                 inv_date, desc, category,
                 pay_method, notes, vendor_name, vendor_email,
                 vendor_addr, w9_fname, w9_data_val,
                 inv_fname, inv_data_val,
                 proof_fname, proof_data_val,
                 cobrand, is_reimb, currency, pay_terms, created_at_val) = row

                breakdown = _parse_json_list(breakdown_str)

                if breakdown and len(breakdown) > 1:
                    # Multi-artist split
                    n = len(breakdown)
                    raw_amounts = [e.get("amount") for e in breakdown]
                    has_any_amount = any(
                        a is not None and str(a).strip() not in ("", "0", "0.0", "0.00")
                        for a in raw_amounts
                    )
                    if has_any_amount:
                        split_amounts = []
                        for a in raw_amounts:
                            try:
                                split_amounts.append(
                                    round(float(str(a).replace(",","").strip()), 2)
                                    if a not in (None, "") else 0.0
                                )
                            except:
                                split_amounts.append(0.0)
                    else:
                        base = float(total_amount or 0)
                        per  = round(base / n, 2)
                        remainder = round(base - per * n, 2)
                        split_amounts = [round(per + remainder, 2)] + [per] * (n - 1)

                    cur.execute(f"""UPDATE expenses
                                     SET status='approved', approved_by={ph}, approved_at={ph},
                                         artist={ph}, song={ph}, amount={ph}, artist_breakdown=NULL
                                    WHERE id={ph}""",
                                (approver, now,
                                 breakdown[0].get("artist",""), breakdown[0].get("song",""),
                                 split_amounts[0], eid))

                    for i, entry in enumerate(breakdown[1:], 1):
                        cur.execute(f"""INSERT INTO expenses (
                            invoice_date, payee, description, category, invoice_number,
                            payment_method, notes, vendor_name, vendor_email, vendor_address,
                            w9_filename, w9_data, invoice_filename, invoice_data,
                            proof_filename, proof_data,
                            cobrand, is_reimbursement, currency, payment_terms, created_at,
                            artist, song, amount,
                            status, approved_by, approved_at, parent_id
                        ) VALUES ({','.join([ph]*28)})""",
                        (inv_date, payee, desc, category, inv_num,
                         pay_method, notes, vendor_name, vendor_email, vendor_addr,
                         w9_fname, w9_data_val, inv_fname, inv_data_val,
                         proof_fname, proof_data_val,
                         cobrand, is_reimb, currency, pay_terms, created_at_val,
                         entry.get("artist",""), entry.get("song",""), split_amounts[i],
                         "approved", approver, now, eid))
                else:
                    # Single artist — standard approval
                    cur.execute(f"""UPDATE expenses SET status='approved', approved_by={ph}, approved_at={ph}
                                    WHERE id={ph}""", (approver, now, eid))

                conn.commit(); conn.close()
                detail_parts = []
                if inv_num: detail_parts.append(f"Invoice #{inv_num}")
                if total_amount: detail_parts.append(f"${total_amount}")
                log_action("invoice_approved", eid, payee,
                           details=" | ".join(detail_parts) if detail_parts else None)
                approved.append(eid)
            except Exception as e:
                failed.append(eid)

        return jsonify({"ok": True, "approved": approved, "failed": failed})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/reject/<int:eid>", methods=["POST"])
@login_required
@admin_required
def reject_entry(eid):
    try:
        reason = (request.json or {}).get("reason", "").strip() if request.is_json else ""
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"""SELECT payee, invoice_number, vendor_name, vendor_email,
                               amount, artist, boom_rep
                        FROM expenses WHERE id={ph}""", (eid,))
        row = cur.fetchone()
        cur.execute(f"DELETE FROM expenses WHERE id={ph}", (eid,))
        conn.commit(); conn.close()
        if row:
            payee, inv_num, vendor_name, vendor_email, amount, artist, boom_rep = row
            log_action("invoice_rejected", eid, payee,
                       details=f"Invoice #{inv_num}" if inv_num else None)
            send_status_email(
                vendor_name or payee, vendor_email, "rejected",
                {"invoice_number": inv_num, "amount": amount, "artist": artist},
                boom_rep=boom_rep, reason=reason or None
            )
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/pending-count")
@login_required
def pending_count():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM expenses WHERE status = 'pending'")
        count = cur.fetchone()[0]; conn.close()
        return jsonify({"count": count})
    except:
        return jsonify({"count": 0})


# ── File viewer routes ────────────────────────────────────────────────────────

@app.route("/invoice/<int:eid>")
@login_required
def view_invoice(eid):
    conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
    cur.execute(f"SELECT invoice_filename, invoice_data FROM expenses WHERE id={ph}", (eid,))
    row = cur.fetchone(); conn.close()
    if not row or not row[1]: return "No invoice file on record.", 404
    return serve_file(row[1], row[0] or f"invoice_{eid}.pdf")

@app.route("/proof/<int:eid>")
@login_required
def view_proof(eid):
    conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
    cur.execute(f"SELECT proof_filename, proof_data FROM expenses WHERE id={ph}", (eid,))
    row = cur.fetchone(); conn.close()
    if not row or not row[1]: return "No proof of payment on record.", 404
    return serve_file(row[1], row[0] or f"proof_{eid}.pdf")

@app.route("/w9/<int:eid>")
@login_required
def view_w9(eid):
    conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
    cur.execute(f"SELECT w9_filename, w9_data FROM expenses WHERE id={ph}", (eid,))
    row = cur.fetchone(); conn.close()
    if not row or not row[1]: return "No W9 on record.", 404
    return serve_file(row[1], row[0] or f"w9_{eid}.pdf")


# ── Listing routes ────────────────────────────────────────────────────────────

@app.route("/recent")
@login_required
def recent():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT invoice_date,payee,amount,artist,song,in_quickbooks,uploaded_to_stem
                       FROM expenses
                       WHERE (status = 'approved' OR status IS NULL) AND deleted IS NOT TRUE
                       ORDER BY id DESC LIMIT 10""")
        rows = cur.fetchall(); conn.close()
        return jsonify([{"date":str(r[0] or ""),"payee":str(r[1] or ""),
                         "amount":r[2],"artist":str(r[3] or ""),"song":str(r[4] or ""),
                         "qb":str(r[5] or ""),"stem":str(r[6] or "")} for r in rows])
    except: return jsonify([])

@app.route("/entries")
@login_required
def entries():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT id,invoice_date,payee,description,category,artist,song,
                              invoice_number,amount,payment_method,payment_date,
                              in_quickbooks,uploaded_to_stem,notes,
                              vendor_submitted,vendor_name,w9_filename,
                              invoice_filename,proof_filename,cobrand,
                              approved_by,approved_at,currency,payment_status,
                              created_at,created_by,vendor_email,is_reimbursement,
                              payment_terms,parent_id,boom_rep
                       FROM expenses
                       WHERE (status = 'approved' OR status IS NULL) AND deleted IS NOT TRUE
                       ORDER BY invoice_date DESC, id DESC""")
        rows = cur.fetchall(); conn.close()
        return jsonify([{"id":r[0],"invoice_date":str(r[1] or ""),"payee":str(r[2] or ""),
                         "description":str(r[3] or ""),"category":str(r[4] or ""),
                         "artist":str(r[5] or ""),"song":str(r[6] or ""),
                         "invoice_number":str(r[7] or ""),"amount":r[8],
                         "payment_method":str(r[9] or ""),"payment_date":str(r[10] or ""),
                         "in_quickbooks":str(r[11] or ""),"uploaded_to_stem":str(r[12] or ""),
                         "notes":str(r[13] or ""),"vendor_submitted":bool(r[14]),
                         "vendor_name":str(r[15] or ""),"w9_filename":str(r[16] or ""),
                         "has_invoice":bool(r[17]),"has_proof":bool(r[18]),
                         "cobrand":bool(r[19]) if r[19] else False,
                         "approved_by":str(r[20] or ""),"approved_at":str(r[21] or ""),
                         "currency":str(r[22] or "USD"),
                         "payment_status":str(r[23] or "Unpaid"),
                         "date_uploaded":str(r[24] or "")[:10],
                         "created_by":str(r[25] or ""),
                         "contact_email":str(r[26] or ""),
                         "is_reimbursement":bool(r[27]),
                         "payment_terms":str(r[28] or ""),
                         "parent_id":r[29],
                         "boom_rep":str(r[30] or "")} for r in rows])
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/payments")
@login_required
def payments():
    return render_template("payments.html",
                           is_admin=is_admin(),
                           current_user=session.get("user_name"))

@app.route("/ledger")
@login_required
def ledger():
    return render_template("ledger.html", categories=CATEGORIES,
                           payment_methods=PAYMENT_METHODS, is_admin=is_admin(),
                           is_john=session.get("user_name")=="John")

@app.route("/danny")
@login_required
@danny_required
def danny():
    return render_template("danny.html", current_user=session.get("user_name"))


@app.route("/danny-entries")
@login_required
@danny_required
def danny_entries():
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind == "pg" else "?"
        cur.execute(f"""SELECT id, invoice_date, payee, description, category, artist, song,
                               invoice_number, amount, payment_method, payment_date,
                               in_quickbooks, uploaded_to_stem, notes,
                               vendor_submitted, vendor_name, w9_filename,
                               invoice_filename, proof_filename, cobrand,
                               approved_by, approved_at, currency, payment_status,
                               created_at, created_by, vendor_email, is_reimbursement,
                               payment_terms, parent_id
                        FROM expenses
                        WHERE (status = 'approved' OR status IS NULL)
                          AND deleted IS NOT TRUE
                          AND LOWER(payment_method) = LOWER({ph})
                        ORDER BY invoice_date DESC, id DESC""", ("PayPal",))
        rows = cur.fetchall(); conn.close()
        return jsonify([{
            "id": r[0], "invoice_date": str(r[1] or ""), "payee": str(r[2] or ""),
            "description": str(r[3] or ""), "category": str(r[4] or ""),
            "artist": str(r[5] or ""), "song": str(r[6] or ""),
            "invoice_number": str(r[7] or ""), "amount": r[8],
            "payment_method": str(r[9] or ""), "payment_date": str(r[10] or ""),
            "in_quickbooks": str(r[11] or ""), "uploaded_to_stem": str(r[12] or ""),
            "notes": str(r[13] or ""), "vendor_submitted": bool(r[14]),
            "vendor_name": str(r[15] or ""), "w9_filename": str(r[16] or ""),
            "has_invoice": bool(r[17]), "has_proof": bool(r[18]),
            "cobrand": bool(r[19]) if r[19] else False,
            "approved_by": str(r[20] or ""), "approved_at": str(r[21] or ""),
            "currency": str(r[22] or "USD"), "payment_status": str(r[23] or "Unpaid"),
            "date_uploaded": str(r[24] or "")[:10], "created_by": str(r[25] or ""),
            "contact_email": str(r[26] or ""), "is_reimbursement": bool(r[27]),
            "payment_terms": str(r[28] or ""), "parent_id": r[29]
        } for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/invoices")
@login_required
def invoices_page():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT id,invoice_date,payee,invoice_number,amount,category,
                              artist,invoice_filename,vendor_submitted,vendor_name,parent_id
                       FROM expenses
                       WHERE invoice_filename IS NOT NULL AND invoice_data IS NOT NULL
                         AND (status = 'approved' OR status IS NULL)
                         AND deleted IS NOT TRUE
                       ORDER BY invoice_date DESC, id DESC""")
        rows = cur.fetchall(); conn.close()
        items = [{"id":r[0],"invoice_date":str(r[1] or ""),"payee":str(r[2] or ""),
                  "invoice_number":str(r[3] or ""),"amount":r[4],"category":str(r[5] or ""),
                  "artist":str(r[6] or ""),"invoice_filename":str(r[7] or ""),
                  "vendor_submitted":bool(r[8]),"vendor_name":str(r[9] or ""),
                  "parent_id":r[10]} for r in rows]
    except Exception as e:
        items = []
    return render_template("invoices.html", items=items, is_admin=is_admin())

@app.route("/w9s")
@login_required
def w9s_page():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT id,created_at,vendor_name,vendor_email,w9_filename,
                              payee,invoice_date,amount
                       FROM expenses
                       WHERE w9_filename IS NOT NULL AND w9_data IS NOT NULL
                         AND (status = 'approved' OR status IS NULL)
                         AND deleted IS NOT TRUE
                       ORDER BY id DESC""")
        rows = cur.fetchall(); conn.close()
        items = [{"id":r[0],"created_at":str(r[1] or ""),"vendor_name":str(r[2] or ""),
                  "vendor_email":str(r[3] or ""),"w9_filename":str(r[4] or ""),
                  "payee":str(r[5] or ""),"invoice_date":str(r[6] or ""),
                  "amount":r[7]} for r in rows]
    except Exception as e:
        items = []
    return render_template("w9s.html", items=items, is_admin=is_admin())


# ── History route ─────────────────────────────────────────────────────────────

@app.route("/history")
@login_required
@history_allowed
def history():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT id, timestamp, user_name, action, entry_id, entry_payee,
                              field, old_value, new_value, details
                       FROM audit_log
                       ORDER BY timestamp DESC
                       LIMIT 2000""")
        rows = cur.fetchall(); conn.close()
        logs = [{"id":r[0], "timestamp":str(r[1] or ""), "user_name":str(r[2] or ""),
                 "action":str(r[3] or ""), "entry_id":r[4],
                 "entry_payee":str(r[5] or ""), "field":str(r[6] or ""),
                 "old_value":str(r[7] or ""), "new_value":str(r[8] or ""),
                 "details":str(r[9] or "")} for r in rows]
    except Exception as e:
        logs = []
    is_john = session.get("user_name") == "John"
    return render_template("history.html", logs=logs, is_admin=is_admin(), is_john=is_john)

@app.route("/clear-history", methods=["POST"])
@login_required
@john_required
def clear_history():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("DELETE FROM audit_log")
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Analytics route ───────────────────────────────────────────────────────────

@app.route("/analytics")
@login_required
@admin_required
def analytics():
    return render_template("analytics.html", is_admin=is_admin())

@app.route("/analytics-data")
@login_required
@admin_required
def analytics_data():
    try:
        # Read optional filter params
        f_date_from  = request.args.get('date_from',  '').strip()
        f_date_to    = request.args.get('date_to',    '').strip()
        f_category   = request.args.get('category',   '').strip()
        f_artist     = request.args.get('artist',     '').strip()
        f_status     = request.args.get('payment_status', '').strip()

        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT invoice_date, category, artist, amount, currency,
                              payment_status, in_quickbooks, uploaded_to_stem
                       FROM expenses
                       WHERE (status = 'approved' OR status IS NULL)
                         AND amount IS NOT NULL AND amount > 0
                         AND deleted IS NOT TRUE
                       ORDER BY invoice_date ASC""")
        rows = cur.fetchall(); conn.close()

        # First pass: collect all unique categories/artists for dropdown options (USD only)
        all_categories = sorted(set(
            (r[1] or "Uncategorized") for r in rows
            if (r[4] or "USD").upper() == "USD"
        ))
        all_artists = sorted(set(
            (r[2] or "No Artist") for r in rows
            if (r[4] or "USD").upper() == "USD"
        ))

        by_category   = {}
        by_artist     = {}
        by_month      = {}
        paid_summary  = {"Paid": 0, "Unpaid": 0, "Partial": 0}
        qb_summary    = {"Yes": 0, "No": 0}
        stem_summary  = {"Yes": 0, "No": 0}
        total_usd     = 0

        for r in rows:
            inv_date, category, artist, amount, currency, pay_status, in_qb, in_stem = r
            amt = float(amount or 0)
            # Only aggregate USD
            if (currency or "USD").upper() != "USD":
                continue

            # Date filter
            if f_date_from or f_date_to:
                try:
                    d = inv_date if hasattr(inv_date, "strftime") else datetime.strptime(str(inv_date)[:10], "%Y-%m-%d").date()
                    d_str = d.strftime("%Y-%m-%d")
                    if f_date_from and d_str < f_date_from: continue
                    if f_date_to   and d_str > f_date_to:   continue
                except: pass

            cat = category or "Uncategorized"
            art = artist or "No Artist"
            ps  = pay_status or "Unpaid"

            # Category / artist / status filters
            if f_category and cat != f_category: continue
            if f_artist   and art != f_artist:   continue
            if f_status   and ps  != f_status:   continue

            total_usd += amt
            by_category[cat] = by_category.get(cat, 0) + amt
            by_artist[art]   = by_artist.get(art, 0) + amt

            # Month bucket
            try:
                if inv_date:
                    d = inv_date if hasattr(inv_date, "strftime") else datetime.strptime(str(inv_date)[:10], "%Y-%m-%d").date()
                    key = d.strftime("%b %Y")
                    by_month[key] = by_month.get(key, 0) + amt
            except: pass

            if ps in paid_summary: paid_summary[ps] += amt

            qb = in_qb or "No"
            if qb in qb_summary: qb_summary[qb] += amt

            stem = in_stem or "No"
            if stem in stem_summary: stem_summary[stem] += amt

        # Sort categories and artists by spend desc, cap at top 10
        by_category = dict(sorted(by_category.items(), key=lambda x: x[1], reverse=True)[:10])
        by_artist   = dict(sorted(by_artist.items(),   key=lambda x: x[1], reverse=True)[:10])

        return jsonify({
            "total_usd":      total_usd,
            "by_category":    by_category,
            "by_artist":      by_artist,
            "by_month":       by_month,
            "paid_summary":   paid_summary,
            "qb_summary":     qb_summary,
            "stem_summary":   stem_summary,
            "all_categories": all_categories,
            "all_artists":    all_artists,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Export routes ─────────────────────────────────────────────────────────────

@app.route("/export")
@login_required
def export_excel():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT invoice_date,payee,description,category,artist,song,
                              invoice_number,amount,payment_method,payment_date,
                              in_quickbooks,qb_entry_date,uploaded_to_stem,stem_upload_date,
                              notes,cobrand,approved_by,approved_at,currency,created_at,created_by
                       FROM expenses
                       WHERE (status = 'approved' OR status IS NULL) AND deleted IS NOT TRUE
                       ORDER BY invoice_date ASC, id ASC""")
        rows = cur.fetchall(); conn.close()
    except Exception as e: return jsonify({"error":str(e)}), 500
    wb = _build_excel(rows); buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"BoomRecords_Expenses_{date.today():%Y-%m-%d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/export-qbo")
@login_required
def export_qbo():
    try:
        conn, kind = get_db(); cur = conn.cursor()
        cur.execute("""SELECT invoice_date,payee,description,category,artist,song,
                              invoice_number,amount,payment_method,payment_date,
                              in_quickbooks,uploaded_to_stem,notes,cobrand,
                              approved_by,approved_at,currency,created_at,created_by
                       FROM expenses
                       WHERE (status = 'approved' OR status IS NULL) AND deleted IS NOT TRUE
                       ORDER BY invoice_date ASC, id ASC""")
        rows = cur.fetchall(); conn.close()
    except Exception as e: return jsonify({"error":str(e)}), 500
    csv_content = _build_csv(rows)
    buf = io.BytesIO(csv_content.encode("utf-8-sig"))  # utf-8-sig adds BOM for Excel compatibility
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"BoomRecords_Expenses_{date.today():%Y-%m-%d}.csv",
                     mimetype="text/csv")

def _build_csv(rows):
    import csv, io as _io
    buf = _io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Date Uploaded", "Added By", "Invoice Date", "Payee / Vendor", "Description", "Category",
        "Artist / Project", "Song", "Invoice #", "Currency", "Amount",
        "Payment Method", "Date Paid", "In QuickBooks?",
        "Uploaded to Stem?", "Notes", "Cobrand", "Approved By", "Approved Date"
    ])
    for r in rows:
        inv_date, payee, desc, category, artist, song, inv_num, amount, \
        pay_method, pay_date, in_qb, in_stem, notes, cobrand, approved_by, approved_at, currency, created_at, created_by = r

        def fmt_d(d):
            if not d: return ""
            try:
                if hasattr(d, 'strftime'): return d.strftime("%m/%d/%Y")
                return datetime.strptime(str(d)[:10], "%Y-%m-%d").strftime("%m/%d/%Y")
            except: return str(d)[:10]

        writer.writerow([
            fmt_d(created_at), created_by or "", fmt_d(inv_date), payee or "", desc or "", category or "",
            artist or "", song or "", inv_num or "",
            currency or "USD",
            f"{float(amount):.2f}" if amount else "",
            pay_method or "", fmt_d(pay_date),
            in_qb or "", in_stem or "", notes or "",
            "Yes" if cobrand else "No",
            approved_by or "", fmt_d(approved_at)
        ])
    return buf.getvalue()

def _build_excel(rows):
    def fill(c): return PatternFill("solid",start_color=c,end_color=c)
    def bdr():
        s=Side(style="thin",color="FFE2E2E2"); return Border(left=s,right=s,top=s,bottom=s)
    wb=Workbook(); ws=wb.active; ws.title="Expense Tracker"
    ws.sheet_view.showGridLines=False; ws.freeze_panes="A3"
    ws.merge_cells("A1:V1"); ws["A1"]="BOOM RECORDS — EXPENSE & RECOUPMENT TRACKER"
    ws["A1"].font=Font(name="Arial",bold=True,size=13,color="FFFFFFFF")
    ws["A1"].fill=fill("FFE31E24"); ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=28
    hdrs=[("A","Date Uploaded",14),("B","Added By",14),("C","Invoice Date",14),
          ("D","Payee / Vendor",22),("E","Description",30),("F","Category",20),
          ("G","Artist / Project",20),("H","Song",20),
          ("I","Invoice #",14),("J","Currency",10),("K","Amount",13),
          ("L","Payment Method",16),("M","Date Paid",14),("N","In QuickBooks?",16),
          ("O","QB Entry Date",14),("P","Uploaded to Stem?",18),("Q","Stem Upload Date",16),
          ("R","Cobrand",10),("S","Notes",30),
          ("T","Approved By",14),("U","Approved Date",14)]
    for col,label,w in hdrs:
        c=ws[f"{col}2"]; c.value=label
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFFFF")
        c.fill=fill("FF333333"); c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=bdr(); ws.column_dimensions[col].width=w
    ws.row_dimensions[2].height=22
    for i,row in enumerate(rows):
        r=i+3; v=list(row)
        # v indices: 0=inv_date,1=payee,2=desc,3=cat,4=artist,5=song,6=inv#,
        #            7=amount,8=pay_method,9=pay_date,10=in_qb,11=qb_date,
        #            12=in_stem,13=stem_date,14=notes,15=cobrand,16=approved_by,17=approved_at,
        #            18=currency,19=created_at,20=created_by
        qb=str(v[10] or ""); stem=str(v[12] or ""); currency_val=str(v[18] or "USD")
        rf=fill("FFD9EAD3" if (qb=="Yes" and stem=="Yes") else "FFFFF2CC" if qb=="No" else "FFFCE5CD" if stem=="No" else ("FFF5F5F5" if r%2==0 else "FFFFFFFF"))
        for col,_,_ in hdrs:
            c=ws[f"{col}{r}"]; c.fill=rf; c.font=Font(name="Arial",size=10)
            c.border=bdr(); c.alignment=Alignment(horizontal="left",vertical="center")
        def dc(col,val,fmt=None,align="left"):
            c=ws[f"{col}{r}"]; c.value=val
            if fmt: c.number_format=fmt
            c.alignment=Alignment(horizontal=align,vertical="center")
        def parse_date(d):
            if not d: return None
            try:
                if hasattr(d, 'date'): return d.date()
                return datetime.strptime(str(d)[:10], "%Y-%m-%d").date()
            except: return str(d)[:10] or None
        amt_fmt = '#,##0.00;(#,##0.00);"-"'
        dc("A", parse_date(v[19]), "MM/DD/YYYY" if v[19] else "", "center")
        dc("B", v[20] or "", "", "center")
        dc("C", parse_date(v[0]), "MM/DD/YYYY" if v[0] else "", "center")
        dc("D",v[1]); dc("E",v[2]); dc("F",v[3],"","center")
        dc("G",v[4]); dc("H",v[5]); dc("I",v[6],"","center")
        dc("J", currency_val, "", "center")
        dc("K", v[7], amt_fmt, "right")
        dc("L",v[8],"","center")
        dc("M", parse_date(v[9]), "MM/DD/YYYY" if v[9] else "", "center")
        dc("N",v[10],"","center")
        dc("O", parse_date(v[11]), "MM/DD/YYYY" if v[11] else "", "center")
        dc("P",v[12],"","center")
        dc("Q", parse_date(v[13]), "MM/DD/YYYY" if v[13] else "", "center")
        dc("R","Yes" if v[15] else "No","","center")
        dc("S",v[14])
        dc("T",v[16] or "","","center")
        approved_at_val = parse_date(v[17])
        dc("U", approved_at_val or "", "MM/DD/YYYY" if approved_at_val else "", "center")
    return wb


# ── Vendor submission ─────────────────────────────────────────────────────────

def _validate_file(file_bytes, mime, prompt):
    """Run a validation prompt against a file and return parsed JSON or None."""
    if not ANTHROPIC_KEY: return None
    b64 = base64.standard_b64encode(file_bytes).decode()
    content = ([{"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64}},
                {"type":"text","text":prompt}]
               if mime == "application/pdf" else
               [{"type":"image","source":{"type":"base64","media_type":mime,"data":b64}},
                {"type":"text","text":prompt}])
    try:
        resp = anthropic.Anthropic(api_key=ANTHROPIC_KEY).messages.create(
            model=MODEL, max_tokens=512, messages=[{"role":"user","content":content}])
        raw = resp.content[0].text.strip()
        if raw.startswith("```"): raw = raw.split("```")[1]; raw = raw[4:] if raw.startswith("json") else raw
        return json.loads(raw.strip())
    except Exception as e:
        print(f"Validation error: {e}"); return None

@app.route("/validate-files", methods=["POST"])
def validate_files():
    """Pre-submission AI check on invoice and W9 files. No login required (public form)."""
    if not ANTHROPIC_KEY:
        return jsonify({"ok": True, "skipped": True})

    result = {"invoice": [], "w9": []}

    if "invoice_file" in request.files and request.files["invoice_file"].filename:
        f = request.files["invoice_file"]
        fb = f.read(); mime = ext_mime(f.filename)
        v = _validate_file(fb, mime, INVOICE_VALIDATE_PROMPT)
        if v:
            issues = list(v.get("issues") or [])
            is_reimb = request.form.get("is_reimbursement") == "yes"
            if not v.get("has_invoice_number"): issues.append("Receipt number is missing." if is_reimb else "Invoice number is missing.")
            if not v.get("has_amount"):         issues.append("Amount is missing or unclear.")
            if not v.get("has_date"):           issues.append("Date is missing.")
            if not v.get("has_payee_name"):     issues.append("Vendor / payee name is missing.")
            if not is_reimb and not v.get("billed_to_boom"): issues.append("Invoice must be billed to Boom.Records LLC.")
            result["invoice"] = list(dict.fromkeys(issues))  # dedupe, preserve order

    if "w9_file" in request.files and request.files["w9_file"].filename:
        f = request.files["w9_file"]
        fb = f.read(); mime = ext_mime(f.filename)
        v = _validate_file(fb, mime, W9_VALIDATE_PROMPT)
        if v:
            issues = []  # ignore AI-generated issues; use only our explicit checks
            if not v.get("is_w9_or_w8"):    issues.append("This doesn't appear to be a W-9 or W-8 form.")
            if not v.get("has_name"):        issues.append("Name field is blank or missing.")
            if not v.get("has_tin_ssn_ein"): issues.append("Tax ID is missing — provide your SSN, EIN, ITIN, or foreign TIN.")
            if not v.get("has_signature"):   issues.append("Signature is missing — the form must be signed.")
            if not v.get("has_signed_date"): issues.append("Signature date is missing.")
            result["w9"] = issues

    return jsonify({"ok": True, "issues": result})

@app.route("/check-invoice")
def check_invoice_dup():
    number = request.args.get("number", "").strip()
    if not number or len(number) < 1:
        return jsonify({"duplicate": False})
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"SELECT id, payee FROM expenses WHERE LOWER(invoice_number)=LOWER({ph}) AND deleted IS NOT TRUE LIMIT 1", (number,))
        row = cur.fetchone(); conn.close()
        if row:
            return jsonify({"duplicate": True, "payee": str(row[1] or "")})
        return jsonify({"duplicate": False})
    except Exception as e:
        return jsonify({"duplicate": False})

@app.route("/check-w9")
def check_w9():
    name = request.args.get("name", "").strip()
    if not name or len(name) < 2:
        return jsonify({"has_w9": False})
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        if kind == "pg":
            cur.execute(f"SELECT id FROM expenses WHERE LOWER(payee)=LOWER({ph}) AND w9_filename IS NOT NULL AND w9_data IS NOT NULL AND deleted IS NOT TRUE LIMIT 1", (name,))
        else:
            cur.execute(f"SELECT id FROM expenses WHERE LOWER(payee)=LOWER({ph}) AND w9_filename IS NOT NULL AND w9_data IS NOT NULL AND deleted IS NOT TRUE LIMIT 1", (name,))
        row = cur.fetchone(); conn.close()
        return jsonify({"has_w9": bool(row)})
    except Exception as e:
        return jsonify({"has_w9": False})

@app.route("/submit", methods=["GET"])
def submit_form():
    return render_template("submit.html", categories=CATEGORIES)

@app.route("/submit", methods=["POST"])
def submit_invoice():
    vendor_name      = request.form.get("vendor_name","").strip()
    vendor_email     = request.form.get("vendor_email","").strip()
    vendor_address   = request.form.get("vendor_address","").strip()
    vendor_artist    = request.form.get("artist","").strip()
    vendor_song      = request.form.get("song","").strip()
    vendor_category  = request.form.get("category","").strip()
    vendor_payment   = request.form.get("payment_preference","").strip()
    vendor_inv_num   = request.form.get("invoice_number_hint","").strip()
    cobrand          = request.form.get("cobrand") == "yes"
    notes            = request.form.get("notes","").strip()
    is_reimbursement = request.form.get("is_reimbursement") == "yes"
    artist_breakdown = request.form.get("artist_breakdown","").strip() or None
    boom_rep         = request.form.get("boom_rep","").strip() or None

    def err(msg):
        return render_template("submit.html", error=msg, categories=CATEGORIES)

    if not vendor_name:    return err("Please enter your legal / government name.")
    if not vendor_email:   return err("Please enter your email address.")
    if not vendor_address: return err("Please enter your mailing address.")
    if not vendor_inv_num: return err("Please enter your invoice number.")
    if not vendor_payment: return err("Please select your preferred payment method.")
    if not vendor_artist:  return err("Please enter the artist or project name.")
    if not vendor_category: return err("Please select a category.")
    if "file" not in request.files or not request.files["file"].filename:
        return err("Please upload your invoice file.")

    # W9 not required for reimbursements
    w9_on_file = False
    if not is_reimbursement:
        try:
            conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
            cur.execute(f"SELECT id FROM expenses WHERE LOWER(payee)=LOWER({ph}) AND w9_filename IS NOT NULL AND w9_data IS NOT NULL AND deleted IS NOT TRUE LIMIT 1", (vendor_name,))
            w9_on_file = bool(cur.fetchone()); conn.close()
        except: pass

    has_new_w9 = "w9_file" in request.files and request.files["w9_file"].filename
    if not is_reimbursement and not has_new_w9 and not w9_on_file:
        return err("Please upload your W9 or W8 form.")
    if is_reimbursement and ("receipt_file" not in request.files or not request.files["receipt_file"].filename):
        return err("Please attach your supporting receipt.")

    file = request.files["file"]; file_bytes = file.read()
    inv_fname = file.filename
    mime = ext_mime(inv_fname)
    fields = extract_fields(file_bytes, mime)
    if not fields.get("payee"): fields["payee"] = vendor_name
    # Use vendor-provided values (override AI extraction)
    fields["category"] = vendor_category
    if vendor_inv_num: fields["invoice_number"] = vendor_inv_num
    if vendor_payment: fields["payment_method"] = vendor_payment
    inv_b64 = base64.b64encode(file_bytes).decode()

    if has_new_w9:
        w9_file = request.files["w9_file"]
        w9_fname = w9_file.filename
        w9_b64 = base64.b64encode(w9_file.read()).decode()
    else:
        w9_fname = None
        w9_b64 = None

    # Capture supporting receipt for reimbursements
    has_receipt = is_reimbursement and "receipt_file" in request.files and request.files["receipt_file"].filename
    if has_receipt:
        receipt_file = request.files["receipt_file"]
        receipt_fname = receipt_file.filename
        receipt_b64 = base64.b64encode(receipt_file.read()).decode()
    else:
        receipt_fname = None
        receipt_b64 = None

    unknowns = get_unknowns(fields)
    row = (parse_date(fields.get("invoice_date")), fields.get("payee",""),
           fields.get("description",""), vendor_category,
           vendor_artist, vendor_song, fields.get("invoice_number",""),
           parse_amount(fields.get("amount",0)),
           fields.get("payment_method",""), None, "No", None, "No", None, notes,
           True, vendor_name, vendor_email, vendor_address,
           w9_fname, w9_b64, inv_fname, inv_b64, receipt_fname, receipt_b64, "pending", cobrand, is_reimbursement,
           artist_breakdown, boom_rep)
    try:
        conn, kind = get_db(); cur = conn.cursor(); ph = "%s" if kind=="pg" else "?"
        cur.execute(f"""INSERT INTO expenses (invoice_date,payee,description,category,
            artist,song,invoice_number,amount,payment_method,payment_date,in_quickbooks,
            qb_entry_date,uploaded_to_stem,stem_upload_date,notes,vendor_submitted,
            vendor_name,vendor_email,vendor_address,w9_filename,w9_data,invoice_filename,invoice_data,
            proof_filename,proof_data,status,cobrand,is_reimbursement,artist_breakdown,boom_rep)
            VALUES ({','.join([ph]*30)})""", row)
        conn.commit(); conn.close()
    except Exception as e:
        return err(f"Submission failed: {e}")

    fields["boom_rep"] = boom_rep
    send_vendor_email(vendor_name, vendor_email, fields, unknowns, w9_fname, is_reimbursement=is_reimbursement)
    return render_template("submit_success.html", vendor_name=vendor_name)

@app.route("/vendors")
@login_required
def vendors_page():
    vendors = []
    conn = None
    try:
        conn, kind = get_db()
        cur = conn.cursor()
        if kind == "pg":
            cur.execute("""
                SELECT payee,
                       COUNT(*) AS invoice_count,
                       SUM(amount) AS total_spent,
                       MAX(invoice_date) AS latest_date,
                       MAX(vendor_email) AS vendor_email,
                       MAX(CASE WHEN w9_filename IS NOT NULL THEN 1 ELSE 0 END) AS has_w9,
                       STRING_AGG(DISTINCT payment_method, ', ') AS methods
                FROM expenses
                WHERE deleted IS NOT TRUE
                  AND payee IS NOT NULL
                  AND payee <> ''
                GROUP BY payee
                ORDER BY total_spent DESC NULLS LAST
            """)
        else:
            cur.execute("""
                SELECT payee,
                       COUNT(*) AS invoice_count,
                       SUM(amount) AS total_spent,
                       MAX(invoice_date) AS latest_date,
                       MAX(vendor_email) AS vendor_email,
                       MAX(CASE WHEN w9_filename IS NOT NULL THEN 1 ELSE 0 END) AS has_w9,
                       '' AS methods
                FROM expenses
                WHERE deleted IS NOT TRUE
                  AND payee IS NOT NULL
                  AND payee <> ''
                GROUP BY payee
                ORDER BY total_spent DESC
            """)
        rows = cur.fetchall()
        vendors = [{
            "payee":     r[0],
            "count":     int(r[1] or 0),
            "total":     float(r[2] or 0),
            "total_fmt": "${:,.0f}".format(float(r[2] or 0)),
            "latest":    str(r[3] or ""),
            "email":     str(r[4] or ""),
            "has_w9":    bool(r[5]),
            "methods":   str(r[6] or ""),
        } for r in rows]
    except Exception as e:
        app.logger.error("vendors_page error: %s", e, exc_info=True)
        vendors = []
    finally:
        if conn:
            try: conn.close()
            except: pass
    return render_template("vendors.html", vendors=vendors, is_admin=is_admin())

@app.route("/vendor/<payee>")
@login_required
def vendor_profile(payee):
    """Vendor profile page showing all approved expenses for a given payee."""
    try:
        conn, kind = get_db()
        cur = conn.cursor()
        ph = "%s" if kind == "pg" else "?"

        # Query all approved expenses for this payee (case-insensitive)
        cur.execute(f"""SELECT id, invoice_date, payee, vendor_name, vendor_email,
                               artist, song, amount, payment_method, payment_status,
                               w9_filename, invoice_filename
                        FROM expenses
                        WHERE (status = 'approved' OR status IS NULL)
                          AND deleted IS NOT TRUE
                          AND LOWER(payee) = LOWER({ph})
                        ORDER BY invoice_date DESC""", (payee,))
        rows = cur.fetchall()
        conn.close()

        if not rows:
            return render_template("vendor.html",
                                 vendor_name=payee,
                                 entries=[],
                                 total_spent=0,
                                 count=0,
                                 latest_w9=None,
                                 vendor_email=None,
                                 payment_methods=[])

        # Extract vendor info from first row
        vendor_email = rows[0][4] or ""
        w9_files = [r[10] for r in rows if r[10]]
        latest_w9 = w9_files[0] if w9_files else None

        # Calculate totals and collect methods
        total_spent = sum(r[7] or 0 for r in rows)
        payment_methods = list(set(r[8] for r in rows if r[8]))

        # Build entries list
        entries = []
        for row in rows:
            entries.append({
                "id": row[0],
                "invoice_date": str(row[1] or ""),
                "artist": row[5] or "",
                "song": row[6] or "",
                "amount": row[7],
                "payment_method": row[8] or "",
                "payment_status": row[9] or "Unpaid",
                "invoice_filename": row[11] or None
            })

        return render_template("vendor.html",
                             vendor_name=payee,
                             entries=entries,
                             total_spent=total_spent,
                             count=len(rows),
                             latest_w9=latest_w9,
                             vendor_email=vendor_email,
                             payment_methods=payment_methods)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/calendar")
@login_required
def calendar():
    """Calendar page showing due dates for unpaid invoices with payment terms."""
    return render_template("calendar.html",
                         is_admin=is_admin(),
                         current_user=session.get("user_name"))

@app.route("/1099")
@login_required
def summary_1099():
    year = request.args.get("year", "2025")
    try:
        year_int = int(year)
    except:
        year_int = 2025

    rows = []
    conn = None
    try:
        conn, kind = get_db()
        cur = conn.cursor()
        # Cast invoice_date to get year — works in both PG and SQLite
        if kind == "pg":
            cur.execute("""
                SELECT payee,
                       MAX(vendor_email)  AS email,
                       SUM(amount)        AS total,
                       COUNT(*)           AS invoices,
                       MAX(CASE WHEN w9_filename IS NOT NULL THEN 1 ELSE 0 END) AS has_w9,
                       STRING_AGG(DISTINCT payment_method, ', ') AS methods
                FROM expenses
                WHERE deleted IS NOT TRUE
                  AND payee IS NOT NULL AND payee <> ''
                  AND EXTRACT(YEAR FROM invoice_date::date) = %s
                GROUP BY payee
                ORDER BY total DESC NULLS LAST
            """, (year_int,))
        else:
            cur.execute("""
                SELECT payee,
                       MAX(vendor_email)  AS email,
                       SUM(amount)        AS total,
                       COUNT(*)           AS invoices,
                       MAX(CASE WHEN w9_filename IS NOT NULL THEN 1 ELSE 0 END) AS has_w9,
                       '' AS methods
                FROM expenses
                WHERE deleted IS NOT TRUE
                  AND payee IS NOT NULL AND payee <> ''
                  AND strftime('%Y', invoice_date) = ?
                GROUP BY payee
                ORDER BY total DESC
            """, (str(year_int),))
        for r in cur.fetchall():
            total = float(r[2] or 0)
            rows.append({
                "payee":     r[0],
                "email":     str(r[1] or ""),
                "total":     total,
                "total_fmt": "${:,.2f}".format(total),
                "invoices":  int(r[3] or 0),
                "has_w9":    bool(r[4]),
                "methods":   str(r[5] or ""),
                "needs_1099": total >= 2000,
            })
    except Exception as e:
        app.logger.error("1099 route error: %s", e, exc_info=True)
        rows = []
    finally:
        if conn:
            try: conn.close()
            except: pass

    # Available years: 2023–current
    import datetime
    current_year = datetime.datetime.now().year
    years = list(range(current_year, 2022, -1))

    return render_template("1099.html",
                           rows=rows,
                           year=year_int,
                           years=years,
                           threshold=2000,
                           is_admin=is_admin())

@app.route("/1099/export")
@login_required
def export_1099():
    year = request.args.get("year", "2025")
    try: year_int = int(year)
    except: year_int = 2025
    rows = []
    conn = None
    try:
        conn, kind = get_db(); cur = conn.cursor()
        if kind == "pg":
            cur.execute("""
                SELECT payee, MAX(vendor_email), SUM(amount), COUNT(*),
                       MAX(CASE WHEN w9_filename IS NOT NULL THEN 1 ELSE 0 END),
                       STRING_AGG(DISTINCT payment_method, ', ')
                FROM expenses
                WHERE deleted IS NOT TRUE AND payee IS NOT NULL AND payee <> ''
                  AND EXTRACT(YEAR FROM invoice_date::date) = %s
                GROUP BY payee ORDER BY SUM(amount) DESC NULLS LAST
            """, (year_int,))
        else:
            cur.execute("""
                SELECT payee, MAX(vendor_email), SUM(amount), COUNT(*),
                       MAX(CASE WHEN w9_filename IS NOT NULL THEN 1 ELSE 0 END), ''
                FROM expenses
                WHERE deleted IS NOT TRUE AND payee IS NOT NULL AND payee <> ''
                  AND strftime('%Y', invoice_date) = ?
                GROUP BY payee ORDER BY SUM(amount) DESC
            """, (str(year_int),))
        rows = cur.fetchall()
    except Exception as e:
        app.logger.error("1099 export error: %s", e, exc_info=True)
    finally:
        if conn:
            try: conn.close()
            except: pass
    import csv, io
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Vendor", "Email", "Total Paid", "Invoices", "W9 on File", "Payment Methods", "1099 Required (>=$2000)"])
    for r in rows:
        total = float(r[2] or 0)
        w.writerow([r[0], r[1] or "", "${:.2f}".format(total), r[3],
                    "Yes" if r[4] else "No", r[5] or "",
                    "Yes" if total >= 2000 else "No"])
    output = out.getvalue()
    return output, 200, {
        "Content-Type": "text/csv",
        "Content-Disposition": f"attachment; filename=1099-summary-{year_int}.csv"
    }

@app.route("/status")
def status(): return jsonify({"ok":True})

if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5100))
    print(f"\n  Boom.Records  →  http://localhost:{port}\n")
    app.run(debug=False, host="0.0.0.0", port=port)
